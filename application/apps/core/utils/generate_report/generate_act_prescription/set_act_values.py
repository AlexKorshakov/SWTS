import datetime
import os
from openpyxl.drawing.image import Image
from xlsxwriter.worksheet import Worksheet

from pandas import DataFrame

from apps.core.database.db_utils import db_get_data_dict_from_table_with_id
from apps.core.utils.generate_report.generate_act_prescription.set_act_alignment import set_act_alignment
from apps.core.utils.generate_report.sheet_formatting.set_font import set_report_font, sets_report_font
from apps.core.utils.generate_report.generate_act_prescription.set_act_frame_border import set_range_border
from apps.core.utils.generate_report.sheet_formatting.set_merge_cells import set_merge_cells
from apps.core.utils.generate_report.sheet_formatting.set_row_dimensions import set_row_dimensions, \
    set_automatic_row_dimensions
from apps.core.utils.img_processor.insert_img import image_preparation, insert_images
from loader import logger


async def set_act_violation_values(worksheet: Worksheet, dataframe: DataFrame, workbook, full_act_path):
    """Заполнение акта значениями из dataframe. Сортировка по main_location_id и sub_locations_id

    :param full_act_path:
    :param workbook:
    :param worksheet: страница для заполнения
    :param dataframe: dataframe с данными нарушений
    :return: None
    """

    # берём уникальные значения main_location_id
    unique_main_locations_ids: list = dataframe.main_location_id.unique().tolist()
    row_number: int = 0
    item_row_value: int = 0

    for main_location_id in unique_main_locations_ids:
        new_header_row = True

        # разделяем dataframe по каждому уникальному значению
        main_location_values: dataframe = dataframe.loc[dataframe['main_location_id'] == main_location_id]
        if main_location_values.empty:
            continue

        # разделяем по sub_location
        unique_sub_locations_ids: list = main_location_values.sub_location_id.unique().tolist()

        for sub_locations_id in unique_sub_locations_ids:

            new_header_row = True
            sub_locations_values: dataframe = main_location_values.loc[dataframe['sub_location_id'] == sub_locations_id]
            if sub_locations_values.empty:
                continue

            # итерируемся по dataframe как по tuple
            for row in sub_locations_values.itertuples(index=False):
                item_row_value += 1
                row_number = await set_act_worksheet_cell_value(
                    worksheet, row, row_number, new_header_row, workbook, full_act_path, item_row_value
                )
                new_header_row = False
                row_number += 1

    return row_number


async def set_act_worksheet_cell_value(worksheet: Worksheet, violation_values: DataFrame, row_number: int,
                                       new_header_row: bool,
                                       workbook, full_act_path: str, item_row_value: int):
    """Заполнение тела акта каждым пунктом

        :param item_row_value:
        :param full_act_path:
        :param workbook:
        :param new_header_row: bool  Требуется ли новый заголовок?
        :param row_number:
        :param violation_values: DataFrame
        :param worksheet: страница для заполнения
        :return
        """

    if row_number == 0:
        logger.debug(violation_values)

        await set_single_violation(worksheet, violation_values)
        workbook.save(full_act_path)
        return row_number

    if new_header_row:
        logger.debug(f'{new_header_row= }')

        await set_value_title(worksheet, violation_values, row_number)
        workbook.save(full_act_path)
        row_number += 1

        await set_act_violation(worksheet, violation_values, row_number=row_number, item_row_value=item_row_value)
        workbook.save(full_act_path)
        return row_number

    if not new_header_row:
        logger.debug(f'{new_header_row= }')

        await set_act_violation(worksheet, violation_values, row_number=row_number, item_row_value=item_row_value)
        workbook.save(full_act_path)
        return row_number


async def set_act_violation(worksheet: Worksheet, violation_values, row_number: int, item_row_value: int):
    """

    :param item_row_value:
    :param row_number:
    :param worksheet:
    :param violation_values:
    :return:
    """

    row_value = 28 + row_number

    worksheet.cell(row=row_value, column=2, value=item_row_value)
    normative_document: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_normativedocuments',
        post_id=violation_values.normative_documents_id
    )

    if violation_values.description:
        worksheet.cell(row=row_value, column=3, value=violation_values.description)
        # worksheet.cell(row=row_value, column=15, value=violation_values.description)
    else:
        worksheet.cell(row=row_value, column=3, value=normative_document['title'])
        # worksheet.cell(row=row_value, column=15, value=normative_document['title'])

    worksheet.cell(row=row_value, column=6, value=normative_document['normative'])
    # worksheet.cell(row=row_value, column=16, value=normative_document['normative'])

    if violation_values.comment and violation_values.comment not in [None, '', ' ', '.', '*', '/']:
        worksheet.cell(row=row_value, column=9, value=violation_values.comment)
        # worksheet.cell(row=row_value, column=17, value=violation_values.comment)
    else:
        worksheet.cell(row=row_value, column=9, value=normative_document['procedure'])
        # worksheet.cell(row=row_value, column=17, value=normative_document['procedure'])

    elimination_time: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_eliminationtime',
        post_id=violation_values.elimination_time_id
    )
    elimination_data = (datetime.datetime.strptime(violation_values.created_at, '%Y-%m-%d')
                        + datetime.timedelta(days=elimination_time['days'])).strftime('%d.%m.%Y')

    worksheet.cell(row=row_value, column=12, value=elimination_data)

    merged_cell = [
        f'C{row_value}:E{row_value}',
        f'F{row_value}:H{row_value}',
        f'I{row_value}:K{row_value}',
    ]

    for item in merged_cell:
        await set_merge_cells(worksheet, merged_cell=item)

    act_range_border = [
        (f'B{row_value}:L{row_value}', False)
    ]
    for item in act_range_border:
        await set_range_border(worksheet, cell_range=item[0], border=item[1])

    row_dimensions = [
        [f'{row_value}', '105'],
    ]
    for item in row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    report_font = [
        (f'B{row_value}:L{row_value}', 'Times New Roman', 11, 'center', 'center'),
        (f'O{row_value}:Q{row_value}', 'Times New Roman', 11, 'center', 'center')
    ]

    elimination_time: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_eliminationtime',
        post_id=violation_values.elimination_time_id
    )
    elimination_data = (datetime.datetime.strptime(violation_values.created_at, '%Y-%m-%d')
                        + datetime.timedelta(days=elimination_time['days'])).strftime('%d.%m.%Y')

    worksheet.cell(row=row_value, column=12, value=elimination_data)

    for item_range in report_font:
        await set_report_font(worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1])

    for item_range in report_font:
        await set_act_alignment(worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4])

    await set_automatic_row_dimensions(worksheet, row_number=row_value, row_value=violation_values)

    return row_value


async def set_act_photographic_materials_values(worksheet: Worksheet, row_value):
    """

    :param row_value:
    :param worksheet:
    """

    values = [
        {'coordinate': 'K59', 'value': 'Приложение 1',
         'row': f'{59 + row_value}', 'column': '12'},
        {'coordinate': 'K60', 'value': 'к 1 части Акта-Предписания',
         'row': f'{60 + row_value}', 'column': '12'},
        {'coordinate': 'B62', 'value': 'Фото 1',
         'row': f'{62 + row_value}', 'column': '2'},
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column']), value=str(val['value']))

        except Exception as err:
            logger.error(f"set_act_photographic_materials_values {repr(err)}")
            return None


async def set_act_photographic_materials(worksheet: Worksheet, img_params: dict):
    """Вставка фото нарушения на страницу акта

    :param img_params:
    :param worksheet:
    :return:
    """

    if not img_params:
        logger.error(f'set_act_photographic_materials not found dict img_params')
        return False

    photo_full_name = img_params.get("photo_full_name", None)
    if not photo_full_name:
        logger.error(f'set_act_photographic_materials not found param img_params["photo_full_name"]')
        return False

    if not os.path.isfile(photo_full_name):
        logger.error(f'set_act_photographic_materials photo not found {photo_full_name}')
        return False

    img: Image = Image(photo_full_name)
    img = await image_preparation(img, img_params)
    await insert_images(worksheet, img=img)

    return True


async def set_single_violation(worksheet: Worksheet, violation_values):
    """Заполнение акта из единственного пункта

    :param
    """

    main_location: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_mainlocation',
        post_id=violation_values.main_location_id
    )
    sub_location: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_sublocation',
        post_id=violation_values.sub_location_id
    )

    title: str = f"{main_location['short_title']} ({sub_location['title']})"

    worksheet.cell(row=27, column=2, value=title)
    worksheet.cell(row=28, column=2, value=1)

    normative_document: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_normativedocuments',
        post_id=violation_values.normative_documents_id
    )

    if violation_values.description:
        worksheet.cell(row=28, column=3, value=violation_values.description)
        # worksheet.cell(row=28, column=15, value=violation_values.description)
    else:
        worksheet.cell(row=28, column=3, value=normative_document['title'])
        # worksheet.cell(row=28, column=15, value=normative_document['title'])

    worksheet.cell(row=28, column=6, value=normative_document['normative'])
    # worksheet.cell(row=28, column=16, value=normative_document['normative'])

    if violation_values.comment and violation_values.comment not in ['', ' ', '.', '*', '/']:
        worksheet.cell(row=28, column=9, value=violation_values.comment)
        # worksheet.cell(row=28, column=17, value=violation_values.comment)
    else:
        worksheet.cell(row=28, column=9, value=normative_document['procedure'])
        # worksheet.cell(row=28, column=17, value=normative_document['procedure'])

    elimination_time: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_eliminationtime',
        post_id=violation_values.elimination_time_id
    )
    elimination_data = (datetime.datetime.strptime(violation_values.created_at, '%Y-%m-%d')
                        + datetime.timedelta(days=elimination_time['days'])).strftime('%d.%m.%Y')

    await set_automatic_row_dimensions(worksheet, row_number=28, row_value=violation_values)

    worksheet.cell(row=28, column=12, value=elimination_data)


async def set_value_title(worksheet: Worksheet, violation_values, row_number: int):
    """

    :param row_number:
    :param violation_values:
    :param worksheet:
    :return:
    """

    row_value = 28 + row_number
    main_location: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_mainlocation',
        post_id=violation_values.main_location_id
    )
    sub_location: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_sublocation',
        post_id=violation_values.sub_location_id
    )

    title: str = f"{main_location['short_title']} ({sub_location['title']})"
    worksheet.cell(row=row_value, column=2, value=title)

    merged_cell = [
        f'B{row_value}:L{row_value}',
    ]
    for item in merged_cell:
        await set_merge_cells(worksheet, merged_cell=item)

    act_range_border = [
        (f'B{row_value}:L{row_value}', False)
    ]
    for item in act_range_border:
        await set_range_border(worksheet, cell_range=item[0], border=item[1])

    row_dimensions = [
        [f'{row_value}', '18'],
    ]
    for item in row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    report_font = [
        (f'B{row_value}:L{row_value}', 'Times New Roman', 14, 'center', 'center'),
    ]
    for item_range in report_font:
        await set_report_font(worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1])

    for item_range in report_font:
        await set_act_alignment(worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4])

    report_font = [
        [f'B{row_value}:L{row_value}',
         {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
    ]
    for cell_range in report_font:
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])