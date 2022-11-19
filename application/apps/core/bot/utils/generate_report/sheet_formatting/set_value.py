import datetime
import os
from math import ceil
from openpyxl.drawing.image import Image
from pprint import pprint
from xlsxwriter.worksheet import Worksheet

from pandas import DataFrame

from app import MyBot
from apps.core.bot.data.category import ELIMINATION_TIME, get_data_list
from apps.core.bot.data.report_data import headlines_data
from apps.core.bot.database.DataBase import DataBase
from apps.core.bot.messages.messages import Messages
from apps.core.bot.utils.generate_report.get_file_list import get_registration_json_file_list
from apps.core.bot.utils.generate_report.settings_mip_report import CATEGORY_LIST_VALUES
from apps.core.bot.utils.generate_report.sheet_formatting.set_alignment import set_mip_alignment, set_act_alignment
from apps.core.bot.utils.generate_report.sheet_formatting.set_font import sets_report_font, set_report_font
from apps.core.bot.utils.generate_report.sheet_formatting.set_frame_border import set_range_border, set_act_range_border
from apps.core.bot.utils.generate_report.sheet_formatting.set_merge_cells import set_merge_cells
from apps.core.bot.utils.generate_report.sheet_formatting.set_row_dimensions import set_row_dimensions
from apps.core.bot.utils.img_processor.insert_img import image_preparation, insert_images
from apps.core.bot.utils.json_worker.read_json_file import read_json_file
from loader import logger

not_found: str = 'не выявлено'
not_tested: str = 'не проверялось'

check_mark_true: str = 'V'
check_mark_false: str = '□'

start_photo_row: int = 52
photo_column: str = 'C'
description_column: int = 7
photo_height: int = 400


async def set_headlines_data_values(chat_id):
    """Формирование заголовков отчета

    :param chat_id: id пользователя для которого заполняется отчет
    :return:
    """

    be_away: str = 'отсутствовал'

    if not headlines_data:
        # Руководитель строительства
        headlines_data['construction_manager'] = be_away
        # Инженер СК
        headlines_data['building_control_engineer'] = be_away
        # Подрядчик
        headlines_data['general_contractor'] = ''
        # Подрядчики
        headlines_data['contractors'] = ''
        # Субподрядчик
        headlines_data['subcontractor'] = ''
        # Проект
        headlines_data['name_location'] = ''
        # Вид обхода
        headlines_data['linear_bypass'] = 'Первичный'
        # Дата
        headlines_data['date_linear_bypass'] = ''
        # Представитель подрядчика
        headlines_data['contractor_representative'] = be_away
        # Представитель субподрядчика
        headlines_data['subcontractor_representative'] = be_away

    headlines_data["day"] = (datetime.datetime.now()).strftime("%d")
    headlines_data["year"] = (datetime.datetime.now()).strftime("%Y")

    # if not user_data:
    registration_file_list: list = await get_registration_json_file_list(chat_id=chat_id)
    if not registration_file_list:
        logger.warning(Messages.Error.registration_file_list_not_found)
        await MyBot.bot.send_message(chat_id=chat_id, text=Messages.Error.file_list_not_found)
        return

    registration_data: dict = await read_json_file(registration_file_list)

    headlines_data['function'] = registration_data.get("function")
    headlines_data['name'] = registration_data.get("name")
    headlines_data['name_location'] = registration_data.get("name_location")
    headlines_data['phone_number'] = registration_data.get("phone_number")
    headlines_data['work_shift'] = registration_data.get("work_shift")

    date_now = datetime.datetime.now().strftime("%d.%m.%Y")
    date_then = datetime.datetime.now() - datetime.timedelta(days=1)
    date_then = date_then.strftime("%d.%m.%Y")

    if headlines_data.get("work_shift").lower() == "дневная смена":

        headlines_data['work_shift'] = 'дневной смены'
        headlines_data['custom_date'] = f"{date_now}"

    else:

        headlines_data['work_shift'] = 'о ночной смене'
        headlines_data['custom_date'] = f"{date_then} - {date_now}"


async def set_report_body_values(worksheet: Worksheet):
    """Заполнение первоначальных данных отчета

    :param worksheet: страница для заполнения
    :return:
    """
    values = [
        {"coordinate": "C2", "value": "МОСИНЖПРОЕКТ", "row": "2", "column": "3"},
        # {"coordinate": "D2", "value": "Отчет о ночной смены", "row": "2", "column": "4"},
        {"coordinate": "C3", "value": "ЛО-МИП-УОТиПБ--", "row": "3", "column": "3"},
        {"coordinate": "D3", "value": "Значение", "row": "3", "column": "4"},
        {"coordinate": "F3", "value": "Примечание", "row": "3", "column": "6"},
        {"coordinate": "C4", "value": "Общая информация", "row": "4", "column": "3"},
        {"coordinate": "C5", "value": "Обход", "row": "5", "column": "3"},
        # {"coordinate": "D5", "value": "Первичный", "row": "5", "column": "4"},
        {"coordinate": "C6", "value": "Дата", "row": "6", "column": "3"},
        {"coordinate": "C7", "value": "Подрядчик", "row": "7", "column": "3"},
        # {"coordinate": "D7", "value": "стм.", "row": "7", "column": "4"},
        {"coordinate": "C8", "value": "Субподрядчик", "row": "8", "column": "3"},
        {"coordinate": "C9", "value": "Проект", "row": "9", "column": "3"},
        # {"coordinate": "D9", "value": "", "row": "9", "column": "4"},
        {"coordinate": "C10", "value": "Комиссия", "row": "10", "column": "3"},
        {"coordinate": "D10", "value": "Функция", "row": "10", "column": "4"},
        {"coordinate": "F10", "value": "ФИО", "row": "10", "column": "6"},
        {"coordinate": "D11", "value": "□", "row": "11", "column": "4"},
        {"coordinate": "E11", "value": "Инспектирующие", "row": "11", "column": "5"},
        {"coordinate": "D12", "value": f"{check_mark_true}", "row": "12", "column": "4"},
        {"coordinate": "E12", "value": "Руководитель строительства", "row": "12", "column": "5"},
        # {"coordinate": "F12", "value": f"{be_away}", "row": "12", "column": "6"},
        {"coordinate": "D13", "value": f"{check_mark_true}", "row": "13", "column": "4"},
        {"coordinate": "E13", "value": "Специалист отдела ПБ", "row": "13", "column": "5"},
        {"coordinate": "D14", "value": f"{check_mark_true}", "row": "14", "column": "4"},
        {"coordinate": "E14", "value": "Инженер СК", "row": "14", "column": "5"},
        # {"coordinate": "F14", "value": f"{be_away}", "row": "14", "column": "6"},
        {"coordinate": "D15", "value": f"{check_mark_true}", "row": "15", "column": "4"},
        {"coordinate": "E15", "value": "Подрядчик", "row": "15", "column": "5"},
        # {"coordinate": "F15", "value": f"{be_away}", "row": "15", "column": "6"},
        {"coordinate": "D16", "value": f"{check_mark_true}", "row": "16", "column": "4"},
        {"coordinate": "E16", "value": "Субподрядчик", "row": "16", "column": "5"},
        {"coordinate": "C17", "value": "Охрана труда, промышленная безопасность и охрана окружающей среды", "row": "17",
         "column": "3"},
        {"coordinate": "C18", "value": "Наблюдения", "row": "18", "column": "3"},
        {"coordinate": "D18", "value": "Категория несоответствия", "row": "18", "column": "4"},
        {"coordinate": "F18", "value": "№", "row": "18", "column": "6"},
        {"coordinate": "G18", "value": "Несоответствие", "row": "18", "column": "7"},
        {"coordinate": "H18", "value": "Срок", "row": "18", "column": "8"},

        {"coordinate": "D19", "value": f"{check_mark_false}", "row": "19", "column": "4"},
        {"coordinate": "E19", "value": "Документы ОТ и ПБ", "row": "19", "column": "5"},
        {"coordinate": "G19", "value": f"{not_tested}", "row": "19", "column": "7"},

        {"coordinate": "D20", "value": f"{check_mark_false}", "row": "20", "column": "4"},
        {"coordinate": "E20", "value": "Обучение/аттестация/квалификация", "row": "20", "column": "5"},
        {"coordinate": "G20", "value": f"{not_found}", "row": "20", "column": "7"},

        {"coordinate": "D21", "value": f"{check_mark_true}", "row": "21", "column": "4"},
        {"coordinate": "E21", "value": "СИЗ", "row": "21", "column": "5"},
        {"coordinate": "G21", "value": f"{not_found}", "row": "21", "column": "7"},

        {"coordinate": "D22", "value": f"{check_mark_true}", "row": "22", "column": "4"},
        {"coordinate": "E22", "value": "Механизмы и оборудование", "row": "22", "column": "5"},
        {"coordinate": "G22", "value": f"{not_found}", "row": "22", "column": "7"},

        {"coordinate": "D23", "value": f"{check_mark_true}", "row": "23", "column": "4"},
        {"coordinate": "E23", "value": "ТС/Спецтехника", "row": "23", "column": "5"},
        {"coordinate": "G23", "value": f"{not_found}", "row": "23", "column": "7"},

        {"coordinate": "D24", "value": f"{check_mark_true}", "row": "24", "column": "4"},
        {"coordinate": "E24", "value": "Знаки безопасности/ограждения", "row": "24", "column": "5"},
        {"coordinate": "G24", "value": f"{not_found}", "row": "24", "column": "7"},

        {"coordinate": "D25", "value": f"{check_mark_true}", "row": "25", "column": "4"},
        {"coordinate": "E25", "value": "Земляные работы", "row": "25", "column": "5"},
        {"coordinate": "G25", "value": f"{not_found}", "row": "25", "column": "7"},

        {"coordinate": "D26", "value": f"{check_mark_true}", "row": "26", "column": "4"},
        {"coordinate": "E26", "value": "Электробезопасность", "row": "26", "column": "5"},
        {"coordinate": "G26", "value": f"{not_found}", "row": "26", "column": "7"},

        {"coordinate": "D27", "value": f"{check_mark_true}", "row": "27", "column": "4"},
        {"coordinate": "E27", "value": "Бетонные работы", "row": "27", "column": "5"},
        {"coordinate": "G27", "value": f"{not_found}", "row": "27", "column": "7"},

        {"coordinate": "D28", "value": f"{check_mark_true}", "row": "28", "column": "4"},
        {"coordinate": "E28", "value": "ГПМ", "row": "28", "column": "5"},
        {"coordinate": "G28", "value": f"{not_found}", "row": "28", "column": "7"},

        {"coordinate": "D29", "value": f"{check_mark_true}", "row": "29", "column": "4"},
        {"coordinate": "E29", "value": "Замкнутые пространства", "row": "29", "column": "5"},
        {"coordinate": "G29", "value": f"{not_found}", "row": "29", "column": "7"},

        {"coordinate": "D30", "value": f"{check_mark_true}", "row": "30", "column": "4"},
        {"coordinate": "E30", "value": "Ручные инструменты", "row": "30", "column": "5"},
        {"coordinate": "G30", "value": f"{not_found}", "row": "30", "column": "7"},

        {"coordinate": "D31", "value": f"{check_mark_true}", "row": "31", "column": "4"},
        {"coordinate": "E31", "value": "Работы на высоте", "row": "31", "column": "5"},
        {"coordinate": "G31", "value": f"{not_found}", "row": "31", "column": "7"},

        {"coordinate": "D32", "value": f"{check_mark_true}", "row": "32", "column": "4"},
        {"coordinate": "E32", "value": "Огневые работы", "row": "32", "column": "5"},
        {"coordinate": "G32", "value": f"{not_found}", "row": "32", "column": "7"},

        {"coordinate": "D33", "value": f"{check_mark_true}", "row": "33", "column": "4"},
        {"coordinate": "E33", "value": "Оборудование под давлением", "row": "33", "column": "5"},
        {"coordinate": "G33", "value": f"{not_found}", "row": "33", "column": "7"},

        {"coordinate": "D34", "value": f"{check_mark_true}", "row": "34", "column": "4"},
        {"coordinate": "E34", "value": "Пожарная безопасность", "row": "34", "column": "5"},
        {"coordinate": "G34", "value": f"{not_found}", "row": "34", "column": "7"},

        {"coordinate": "D35", "value": f"{check_mark_true}", "row": "35", "column": "4"},
        {"coordinate": "E35", "value": "Первая помощь", "row": "35", "column": "5"},
        {"coordinate": "G35", "value": f"{not_found}", "row": "35", "column": "7"},

        {"coordinate": "D36", "value": f"{check_mark_false}", "row": "36", "column": "4"},
        {"coordinate": "E36", "value": "Химические, биологические факторы", "row": "36", "column": "5"},
        {"coordinate": "G36", "value": f"{not_tested}", "row": "36", "column": "7"},

        {"coordinate": "D37", "value": f"{check_mark_true}", "row": "37", "column": "4"},
        {"coordinate": "E37", "value": "Санитарные требования", "row": "37", "column": "5"},
        {"coordinate": "G37", "value": f"{not_found}", "row": "37", "column": "7"},

        {"coordinate": "D38", "value": f"{check_mark_true}", "row": "38", "column": "4"},
        {"coordinate": "E38", "value": "Складирование", "row": "38", "column": "5"},
        {"coordinate": "G38", "value": f"{not_found}", "row": "38", "column": "7"},

        {"coordinate": "D39", "value": f"{check_mark_true}", "row": "39", "column": "4"},
        {"coordinate": "E39", "value": "Безопасные проходы", "row": "39", "column": "5"},
        {"coordinate": "G39", "value": f"{not_found}", "row": "39", "column": "7"},

        {"coordinate": "D40", "value": f"{check_mark_true}", "row": "40", "column": "4"},
        {"coordinate": "E40", "value": "Отходы", "row": "40", "column": "5"},
        {"coordinate": "G40", "value": f"{not_found}", "row": "40", "column": "7"},

        {"coordinate": "D41", "value": f"{check_mark_true}", "row": "41", "column": "4"},
        {"coordinate": "E41", "value": "Дежурное освещение", "row": "41", "column": "5"},
        {"coordinate": "G41", "value": f"{not_found}", "row": "41", "column": "7"},

        {"coordinate": "D42", "value": f"{check_mark_true}", "row": "42", "column": "4"},
        {"coordinate": "E42", "value": "Другое", "row": "42", "column": "5"},
        {"coordinate": "G42", "value": f"{not_found}", "row": "42", "column": "7"},

        {"coordinate": "C43", "value": "Дополнительная информация", "row": "43", "column": "3"},
        {"coordinate": "C44",
         "value": 'Данное сообщение рассылается Блоком по качеству, охране труда, промышленной безопасности и охране '
                  'окружающей среды АО Мосинжпроект с целью информирования о состоянии площадки, производства и '
                  'документирования строительно-монтажных работ',
         "row": "44", "column": "3"}
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column'])).value = str(val['value'])

            if val["value"] == not_found:
                cell_range = [f"D{val['row']}:H{val['row']}",
                              {"color": "008000", "font_size": "14", "name": "Arial"}]
                await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

        except Exception as err:
            logger.error(f"set_values {repr(err)}")
            return None


async def set_report_header_values(worksheet: Worksheet, dataframe):
    """Заполнение заголовка отчета

    :param worksheet: страница для заполнения
    :return:
    """

    contractors = dataframe.general_contractor.tolist()
    contractors = list(set(list(contractors)))
    contractors_str = ''
    for item in contractors:
        if isinstance(item, float):
            continue
        if item and item != 'Подрядная организация':
            contractors_str = contractors_str + item + ', '
            headlines_data['contractors'] = contractors_str

    values = [
        {"coordinate": "D2", "value": f"Отчет {headlines_data.get('work_shift')} {headlines_data.get('custom_date')}",
         "row": "2", "column": "4"},
        {"coordinate": "C3", "value": f"ЛО-МИП-УОТиПБ-{headlines_data.get('year')}-{headlines_data.get('day')}",
         "row": "3", "column": "3"},
        {"coordinate": "D5", "value": f"{headlines_data.get('linear_bypass')}", "row": "5", "column": "4"},
        {"coordinate": "D6", "value": f"{headlines_data.get('custom_date')}", "row": "6", "column": "4"},

        {"coordinate": "D7", "value": f"{headlines_data.get('contractors')}", "row": "7", "column": "4"},

        {"coordinate": "D8", "value": "", "row": "8", "column": "4"},
        {"coordinate": "D9", "value": f"{headlines_data.get('name_location')}", "row": "9", "column": "4"},
        {"coordinate": "F12", "value": f"{headlines_data.get('construction_manager')}", "row": "12", "column": "6"},

        {"coordinate": "E13", "value": f"{headlines_data.get('function')}", "row": "13", "column": "5"},
        {"coordinate": "E13", "value": f"{headlines_data.get('name')} тел. +{headlines_data.get('phone_number')}",
         "row": "13", "column": "6"},

        {"coordinate": "F14", "value": f"{headlines_data.get('building_control_engineer')}", "row": "14",
         "column": "6"},
        {"coordinate": "F15", "value": f"{headlines_data.get('contractor_representative')}", "row": "15",
         "column": "6"},
        {"coordinate": "F16", "value": f"{headlines_data.get('subcontractor_representative')}", "row": "16",
         "column": "6"},
    ]

    for val in values:
        try:

            worksheet.cell(row=int(val['row']), column=int(val['column']), value=str(val['value']))

        except Exception as err:
            logger.error(f"set_user_values {repr(err)}")
            return None


async def set_report_violation_values(worksheet: Worksheet, dataframe: DataFrame):
    """Заполнение акта значениями из dataframe

    :param worksheet: страница для заполнения
    :param dataframe: dataframe с данными нарушений
    :return: None
    """

    serial_number = 0
    violation_value = []
    for category in get_data_list("CATEGORY"):
        for item in range(1, dataframe.category.size):
            if dataframe.loc[item]["category"] != category:
                continue

            try:
                elimination_time = await get_elimination_time(dataframe, item)
                violation_value.append(
                    {"description": dataframe.loc[item]["description"] + ' \\',
                     "elimination_time": elimination_time + ' \\'}
                )
            except Exception as err:
                logger.error(f"{repr(err)}")
                continue

        if not violation_value:
            continue

        serial_number += 1
        for item in CATEGORY_LIST_VALUES:

            if category != item['value']:
                continue

            await set_worksheet_cell_value(worksheet, item, serial_number, violation_value)
            violation_value = []
            break


async def set_act_violation_values(worksheet: Worksheet, dataframe: DataFrame, workbook, full_act_path):
    """Заполнение акта значениями из dataframe. Сортировка по main_location_id и sub_locations_id

    :param full_act_path:
    :param workbook:
    :param worksheet: страница для заполнения
    :param dataframe: dataframe с данными нарушений
    :return: None
    """

    # берём уникальные значения main_location_id
    unique_main_locations_ids: list = dataframe.main_location_id.unique().tolist()
    row_number: int = 0
    item_row_value: int = 0

    for main_location_id in unique_main_locations_ids:
        new_header_row = True

        # разделяем dataframe по каждому уникальному значению
        main_location_values: dataframe = dataframe.loc[dataframe['main_location_id'] == main_location_id]
        if main_location_values.empty:
            continue

        # разделяем по sub_location
        unique_sub_locations_ids: list = main_location_values.sub_location_id.unique().tolist()

        for sub_locations_id in unique_sub_locations_ids:

            new_header_row = True
            sub_locations_values: dataframe = main_location_values.loc[dataframe['sub_location_id'] == sub_locations_id]
            if sub_locations_values.empty:
                continue

            # итерируемся по dataframe как по tuple
            for row in sub_locations_values.itertuples(index=False):
                item_row_value += 1
                row_number = await set_act_worksheet_cell_value(
                    worksheet, row, row_number, new_header_row, workbook, full_act_path, item_row_value
                )
                new_header_row = False
                row_number += 1

    return row_number


async def get_elimination_time(dataframe, item) -> str:
    """Получить время устранения

    :return: str
    """
    if dataframe.loc[item]["elimination_time"] != ELIMINATION_TIME[0] and \
            dataframe.loc[item]["elimination_time"] != ELIMINATION_TIME[1]:
        add_days = int(dataframe.loc[item]["elimination_time"].split(' ')[0])
        elimination_time = datetime.datetime.now() + datetime.timedelta(days=add_days)

        return str(elimination_time.strftime("%d.%m.%Y"))

    return dataframe.loc[item]["elimination_time"]


async def set_act_worksheet_cell_value(worksheet: Worksheet, violation_values: DataFrame, row_number: int, new_header_row: bool,
                                       workbook, full_act_path: str, item_row_value: int):
    """Заполнение тела акта каждым пунктом

        :param item_row_value:
        :param full_act_path:
        :param workbook:
        :param new_header_row: bool  Требуется ли новый заголовок?
        :param row_number:
        :param violation_values: DataFrame
        :param worksheet: страница для заполнения
        :return
        """
    if row_number == 0:
        logger.debug(violation_values)

        await set_single_violation(worksheet, violation_values)
        workbook.save(full_act_path)
        return row_number

    if new_header_row:
        logger.debug(f'{new_header_row= }')

        await set_value_title(worksheet, violation_values, row_number)
        workbook.save(full_act_path)
        row_number += 1

        await set_act_violation(worksheet, violation_values, row_number=row_number, item_row_value=item_row_value)
        workbook.save(full_act_path)
        return row_number

    if not new_header_row:
        logger.debug(f'{new_header_row= }')

        await set_act_violation(worksheet, violation_values, row_number=row_number, item_row_value=item_row_value)
        workbook.save(full_act_path)
        return row_number


async def set_value_title(worksheet: Worksheet, violation_values, row_number: int):
    """

    :param row_number:
    :param violation_values:
    :param worksheet:
    :return:
    """

    row_value = 28 + row_number
    main_location: dict = DataBase().get_dict_data_from_table_from_id(
        table_name='core_mainlocation',
        id=violation_values.main_location_id
    )
    sub_location: dict = DataBase().get_dict_data_from_table_from_id(
        table_name='core_sublocation',
        id=violation_values.sub_location_id
    )

    title: str = f"{main_location['short_title']} ({sub_location['title']})"
    worksheet.cell(row=row_value, column=2, value=title)

    merged_cell = [
        f'B{row_value}:L{row_value}',
    ]
    for item in merged_cell:
        await set_merge_cells(worksheet, merged_cell=item)

    act_range_border = [
        (f'B{row_value}:L{row_value}', False)
    ]
    for item in act_range_border:
        await set_act_range_border(worksheet, cell_range=item[0], border=item[1])

    row_dimensions = [
        [f'{row_value}', '18'],
    ]
    for item in row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    report_font = [
        (f'B{row_value}:L{row_value}', 'Times New Roman', 14, 'center', 'center'),
    ]
    for item_range in report_font:
        await set_report_font(worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1])

    for item_range in report_font:
        await set_act_alignment(worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4])

    report_font = [
        [f'B{row_value}:L{row_value}',
         {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
    ]
    for cell_range in report_font:
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])


async def set_act_violation(worksheet: Worksheet, violation_values, row_number: int, item_row_value: int):
    """

    :param item_row_value:
    :param row_number:
    :param worksheet:
    :param violation_values:
    :return:
    """

    row_value = 28 + row_number

    worksheet.cell(row=row_value, column=2, value=item_row_value)
    normative_document: dict = DataBase().get_dict_data_from_table_from_id(
        table_name='core_normativedocuments',
        id=violation_values.normative_documents_id
    )

    if violation_values.description:
        worksheet.cell(row=row_value, column=3, value=violation_values.description)
        worksheet.cell(row=row_value, column=15, value=violation_values.description)
    else:
        worksheet.cell(row=row_value, column=3, value=normative_document['title'])
        worksheet.cell(row=row_value, column=15, value=normative_document['title'])

    worksheet.cell(row=row_value, column=6, value=normative_document['normative'])
    worksheet.cell(row=row_value, column=16, value=normative_document['normative'])

    if violation_values.comment and violation_values.comment not in [None, '', ' ', '.', '*', '/']:
        worksheet.cell(row=row_value, column=9, value=violation_values.comment)
        worksheet.cell(row=row_value, column=17, value=violation_values.comment)
    else:
        worksheet.cell(row=row_value, column=9, value=normative_document['procedure'])
        worksheet.cell(row=row_value, column=17, value=normative_document['procedure'])

    elimination_time: dict = DataBase().get_dict_data_from_table_from_id(
        table_name='core_eliminationtime',
        id=violation_values.elimination_time_id
    )
    elimination_data = (datetime.datetime.strptime(violation_values.created_at, '%Y-%m-%d')
                        + datetime.timedelta(days=elimination_time['days'])).strftime('%d.%m.%Y')

    worksheet.cell(row=row_value, column=12, value=elimination_data)

    merged_cell = [
        f'C{row_value}:E{row_value}',
        f'F{row_value}:H{row_value}',
        f'I{row_value}:K{row_value}',
    ]

    for item in merged_cell:
        await set_merge_cells(worksheet, merged_cell=item)

    act_range_border = [
        (f'B{row_value}:L{row_value}', False)
    ]
    for item in act_range_border:
        await set_act_range_border(worksheet, cell_range=item[0], border=item[1])

    row_dimensions = [
        [f'{row_value}', '105'],
    ]

    for item in row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    report_font = [
        (f'B{row_value}:L{row_value}', 'Times New Roman', 11, 'center', 'center'),
    ]
    for item_range in report_font:
        await set_report_font(worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1])

    for item_range in report_font:
        await set_act_alignment(worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4])

    return row_value


async def set_single_violation(worksheet: Worksheet, violation_values):
    """Заполнение акта из единственного пункта

    :param
    """

    main_location: dict = DataBase().get_dict_data_from_table_from_id(
        table_name='core_mainlocation',
        id=violation_values.main_location_id
    )
    sub_location: dict = DataBase().get_dict_data_from_table_from_id(
        table_name='core_sublocation',
        id=violation_values.sub_location_id
    )

    title: str = f"{main_location['short_title']} ({sub_location['title']})"

    worksheet.cell(row=27, column=2, value=title)
    worksheet.cell(row=28, column=2, value=1)

    normative_document: dict = DataBase().get_dict_data_from_table_from_id(
        table_name='core_normativedocuments',
        id=violation_values.normative_documents_id
    )

    if violation_values.description:
        worksheet.cell(row=28, column=3, value=violation_values.description)
        worksheet.cell(row=28, column=15, value=violation_values.description)
    else:
        worksheet.cell(row=28, column=3, value=normative_document['title'])
        worksheet.cell(row=28, column=15, value=normative_document['title'])

    worksheet.cell(row=28, column=6, value=normative_document['normative'])
    worksheet.cell(row=28, column=16, value=normative_document['normative'])

    if violation_values.comment and violation_values.comment not in ['', ' ', '.', '*', '/']:
        worksheet.cell(row=28, column=9, value=violation_values.comment)
        worksheet.cell(row=28, column=17, value=violation_values.comment)
    else:
        worksheet.cell(row=28, column=9, value=normative_document['procedure'])
        worksheet.cell(row=28, column=17, value=normative_document['procedure'])

    elimination_time: dict = DataBase().get_dict_data_from_table_from_id(
        table_name='core_eliminationtime',
        id=violation_values.elimination_time_id
    )
    elimination_data = (datetime.datetime.strptime(violation_values.created_at, '%Y-%m-%d')
                        + datetime.timedelta(days=elimination_time['days'])).strftime('%d.%m.%Y')

    worksheet.cell(row=28, column=12, value=elimination_data)


async def set_worksheet_cell_value(worksheet: Worksheet, item, serial_number, val):
    """

    :param val:
    :param worksheet: страница для заполнения
    :param item:
    :param serial_number:
    :return:
    """
    worksheet.cell(row=int(item['row']), column=6, value=serial_number)
    worksheet.cell(row=int(item['row']), column=7, value="".join([str(i["description"]) + ' \n' for i in val]))
    worksheet.cell(row=int(item['row']), column=8, value="".join([str(i["elimination_time"]) + ' \n' for i in val]))

    cell_ranges = [
        [f"D{item['row']}:F{item['row']}", {"color": "FF0000", "font_size": "14", "name": "Arial"}],
        [f"G{item['row']}:G{item['row']}", {"color": "FF0000", "font_size": "12", "name": "Arial"}],
        [f"H{item['row']}:H{item['row']}", {"color": "FF0000", "font_size": "12", "name": "Arial"}],
    ]
    for cell_range in cell_ranges:
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])
        await set_mip_alignment(worksheet, cell_range[0], horizontal='left', vertical='center')

    worksheet.row_dimensions[int(item['row'])].height = await get_height_for_row(
        worksheet,
        int(item['row']),
        value="".join([str(i["description"]) + ' \n' for i in val]))


factor_of_font_size_to_width = {

    12: {
        "factor": 0.9,  # width / count of symbols at row
        "height": 25
    }
}


async def get_height_for_row(sheet, row_number: int, font_size: int = 12, value=None):
    """Изменение высоты строк в зависимости от содержания

    :param value:
    :param sheet:
    :param row_number:
    :param font_size:
    :return:
    """
    font_params = factor_of_font_size_to_width[font_size]
    row = list(sheet.rows)[row_number]
    height = font_params["height"]

    for index, cell in enumerate(row):
        if index != 6:
            continue

        if hasattr(cell, 'column_letter'):
            words_count_at_one_row = sheet.column_dimensions[cell.column_letter].width / font_params["factor"]

            lines = ceil(len(str(value)) / words_count_at_one_row)
            height = max(height, lines * font_params["height"])

            return height


async def set_photographic_materials_values(worksheet: Worksheet):
    """

    :param worksheet:
    """

    values = [
        {"coordinate": "C50", "value": "Фотофиксация", "row": "50", "column": "3"},
        {"coordinate": "C51", "value": "ФОТО", "row": "51", "column": "3"},
        {"coordinate": "F51", "value": "№", "row": "51", "column": "6"},
        {"coordinate": "C51", "value": "Примечание", "row": "51", "column": "7"},
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column']), value=str(val['value']))

        except Exception as err:
            logger.error(f"set_photographic_materials_values {repr(err)}")
            return None


async def set_act_photographic_materials_values(worksheet: Worksheet, row_value):
    """

    :param row_value:
    :param worksheet:
    """

    values = [
        {'coordinate': 'K59', 'value': 'Приложение 1',
         'row': f'{59 + row_value}', 'column': '12'},
        {'coordinate': 'K60', 'value': 'к 1 части Акта-Предписания',
         'row': f'{60 + row_value}', 'column': '12'},
        {'coordinate': 'B62', 'value': 'Фото 1',
         'row': f'{62 + row_value}', 'column': '2'},
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column']), value=str(val['value']))

        except Exception as err:
            logger.error(f"set_act_photographic_materials_values {repr(err)}")
            return None


async def set_photographic_materials(worksheet: Worksheet, violation_data: str, num_data: int):
    """Вставка фото и описания наоушения га сстраницу отчета
    :param num_data:
    :param worksheet:
    :param violation_data: 
    :return: 
    """

    img_data = await read_json_file(violation_data)

    if not os.path.isfile(img_data['photo_full_name']):
        logger.error("photo not found")
        return False

    merged_cells = [
        f'C{start_photo_row + num_data}:E{start_photo_row + num_data}',
        f'G{start_photo_row + num_data}:H{start_photo_row + num_data}',
    ]

    for merged_cell in merged_cells:
        worksheet.merge_cells(merged_cell)

    photographic_materials_alignment = [
        f'C{start_photo_row + num_data}:H{start_photo_row + num_data}',
    ]

    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_mip_alignment(worksheet, cell_range, horizontal='left', vertical='center')

    photographic_row_dimensions = [
        [f"{start_photo_row + num_data}", 300.0],
    ]

    for item in photographic_row_dimensions:
        worksheet.row_dimensions[int(item[0])].height = float(item[1])

    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_report_font(worksheet, cell_range=cell_range, font_size=14)

    img: Image = Image(img_data['photo_full_name'])

    img_params: dict = {
        "column": photo_column,
        "row": start_photo_row + num_data,
        "height": photo_height,
        "anchor": True,
        "scale": True,
    }

    img = await image_preparation(img, img_params)

    await insert_images(worksheet, img=img)

    if not img_data.get('description'):
        return False

    worksheet.cell(row=start_photo_row + num_data, column=description_column, value=str(img_data['description']))

    for item, cell_border_range in enumerate(photographic_materials_alignment, start=1):
        await set_range_border(worksheet, cell_range=cell_border_range)

    worksheet.print_area = f'$A$1:$I${start_photo_row + num_data + 1}'

    return True


async def set_act_photographic_materials(worksheet: Worksheet, img_params: dict):
    """Вставка фото нарушения на страницу акта

    :param img_params:
    :param worksheet:
    :return:
    """

    if not img_params:
        logger.error(f'set_act_photographic_materials not found dict img_params')
        return False

    photo_full_name = img_params.get("photo_full_name", None)
    if not photo_full_name:
        logger.error(f'set_act_photographic_materials not found param img_params["photo_full_name"]')
        return False

    if not os.path.isfile(photo_full_name):
        logger.error(f'set_act_photographic_materials photo not found {photo_full_name}')
        return False

    img: Image = Image(photo_full_name)

    img = await image_preparation(img, img_params)

    await insert_images(worksheet, img=img)

    return True
