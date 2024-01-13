from __future__ import annotations

import asyncio
import io
import json
import os
import traceback
from datetime import datetime, timedelta
from pathlib import Path
from pprint import pprint

from aiogram import types
from aiogram.dispatcher import FSMContext
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from apps.MyBot import MyBot, bot_send_message, bot_send_document
from apps.core.bot.bot_utils.check_access import check_developer_access
from apps.core.bot.bot_utils.check_user_registration import check_user_access
from apps.core.bot.handlers.s_user_fanc.s_admin_support_paths import (sa_check_path,
                                                                      sa_check_or_create_dir,
                                                                      sa_get_file_path)
from apps.core.bot.keyboards.inline.build_castom_inlinekeyboard import posts_cb
from apps.core.bot.messages.messages import Messages
from apps.core.bot.states import AnswerUserState
from config.config import Udocan_media_path
from loader import logger


@MyBot.dp.callback_query_handler(posts_cb.filter(action=['s_user_get_files']))
async def call_s_user_get_files(call: types.CallbackQuery, callback_data: dict[str, str]):
    """Обработка ответов содержащихся в s_user_get_files
    """
    hse_user_id = call.message.chat.id
    if not await check_user_access(chat_id=hse_user_id):
        logger.error(f'{await fanc_name()} access fail {hse_user_id = }')
        return

    if not await check_developer_access(chat_id=hse_user_id):
        logger.error(f'{await fanc_name()} check_super_user_access fail {hse_user_id = }')
        return

    await bot_send_message(chat_id=hse_user_id, text=Messages.Enter.address)

    # Вызов состояния ожидания текстового ответа от пользователя
    await AnswerUserState.address.set()
    return True


# Сюда приходит ответ с description, state=состояние
@MyBot.dp.message_handler(state=AnswerUserState.address)
async def process_description(message: types.Message = None, state: FSMContext = None,
                              user_id: int | str = None, test_path: str | Path = None):
    """Обработчик состояния address
    """
    current_state = await state.get_state()
    logger.info(f'{await fanc_name()} Cancelling state {current_state}')
    await state.finish()

    hse_user_id = message.chat.id if message else user_id
    if not await check_user_access(chat_id=hse_user_id):
        logger.error(f'{await fanc_name()} access fail {hse_user_id = }')
        return

    if not await check_developer_access(chat_id=hse_user_id):
        logger.error(f'{await fanc_name()} check_super_user_access fail {hse_user_id = }')
        return

    main_path: str | Path = message.text if message else test_path
    if not sa_check_path(main_path):
        logger.error(f'{await fanc_name()} path {main_path} is not exists!')
        await bot_send_message(chat_id=hse_user_id, text=f'path {main_path} is not exists!')
        return

    save_path: str | Path = await sa_get_file_path(Udocan_media_path, 'temp_dir')
    await sa_check_or_create_dir(save_path)

    file_list: list = await get_file_list(main_path)

    # await write_json(file_path=str(await sa_get_file_path(save_path, f'{get_today()} report.json')), data=file_list)

    doc_path = str(await sa_get_file_path(str(save_path), f'{await get_today()} report.xlsx'))

    await create_file_xlsx(file_list, save_path, doc_path)

    await bot_send_document(
        chat_id=hse_user_id,
        doc_path=doc_path,
        caption='Отчет собран',
        calling_fanc_name=await fanc_name()
    )

    try:
        os.remove(doc_path)
    except PermissionError:
        pass


async def get_file_list(main_path: str):
    """"""

    industrial_equipment_file_list: list = await get_files(main_path)
    return industrial_equipment_file_list


async def get_files(main_path: str) -> list:
    """Получение списка файлов c расширением endswith из main_path

    :type main_path: str
    :param main_path: директория для поиска файлов
    """

    suffix_list: list = [
        '.pyc', '.lnk', '.dwg', '.xml', '.css', '.ini', '.jfif', '.example', '.h', '.po', '.pack',
        '.mo', '.md', '.idx', '.map', '.js', '.hash', '.dat', '.afm', '.pb'
    ]

    suffix_dict = {}
    files_list: list = []
    for subdir, dirs, files in os.walk(main_path):

        for num, file in enumerate(files, start=1):

            if not file: continue
            if '~' in file: continue

            filepath = await sa_get_file_path(subdir, file)

            if 'venv' in filepath.absolute().parts: continue
            if 'build' in filepath.absolute().parts: continue
            if '.git' in filepath.absolute().parts: continue
            if '.mypy_cache' in filepath.absolute().parts: continue
            if 'Lib' in filepath.absolute().parts: continue
            if 'source' in filepath.absolute().parts: continue

            if not filepath.suffix: continue
            if filepath.suffix in suffix_list: continue
            if len(filepath.suffix) == 1: continue
            if len(filepath.suffix) > 5: continue

            try:
                if int(filepath.suffix.split('.')[-1]) > 0: continue
            except ValueError as err:
                pass

            if filepath.suffix not in suffix_dict:
                suffix_dict[filepath.suffix] = 1
            else:
                suffix_dict[filepath.suffix] = suffix_dict.get(filepath.suffix) + 1

            main_path_str = str(main_path)
            suffix = filepath.suffix
            file_dir = os.path.join(*(filepath.absolute().parts[-2:-1])) \
                .replace(str(main_path_str), '')
            subdir_path = os.path.join(*(filepath.absolute().parts[:-1])) \
                .replace(main_path_str, '')
            previous_dir_path = os.path.join(*(filepath.absolute().parts[:-2])) \
                .replace(main_path_str, '').replace(os.sep, '\\')
            full_file_path = str(filepath.absolute())

            try:
                size, suf = await siz_of_file(os.path.getsize(filepath))
            except OSError as err:
                size, suf = 0, ''
            try:
                mod_date = datetime.utcfromtimestamp(os.path.getmtime(filepath)).strftime('%d.%m.%Y')
            except OSError as err:
                mod_date = ''
            try:
                mod_time = datetime.utcfromtimestamp(os.path.getmtime(filepath)).strftime('%H.%M')
            except OSError as err:
                mod_time = ''
            try:
                creation_date = datetime.utcfromtimestamp(os.path.getctime(filepath)).strftime('%d.%m.%Y')
            except OSError as err:
                creation_date = ''

            files_list.append(
                {
                    "file": file,
                    "suffix": suffix,
                    "size": size,
                    "suf": suf,
                    "mod_date": mod_date,
                    "mod_time": mod_time,
                    "creation_date": creation_date,
                    "file_dir": file_dir,
                    "subdir": subdir,
                    "subdir_path": subdir_path,
                    "previous_dir_path": previous_dir_path,
                    "full_file_path_path": full_file_path,
                }
            )
    pprint(suffix_dict)

    return files_list


async def siz_of_file(num, suffix='B'):
    for unit in ('', 'K', 'M', 'G', 'T'):
        if abs(num) < 1024.0:
            return f'{num:3.1f}', f'{unit}{suffix}'
        num /= 1024.0
    return f'{num:.1f}', f'{unit}{suffix}'


async def create_file_xlsx(file_list: list, xlsx_path: str, doc_path: str):
    """Создание файла xlsx с со списком файлов по пути xlsx_path

    :param file_list:
    :param xlsx_path:
    :param doc_path:
    :return:
    """
    await sa_check_or_create_dir(xlsx_path)

    if not file_list:
        return False

    workbook, worksheet = await create_xlsx(str(xlsx_path), doc_path)
    if not workbook or not worksheet:
        return False

    worksheet.cell(row=1, column=1, value='row')
    for col, key in enumerate(file_list[0].keys(), start=2):
        worksheet.cell(row=1, column=col, value=key)

    for row, item in enumerate(file_list, start=2):
        try:
            worksheet.cell(row=row, column=1, value=row - 1)
            for col, key in enumerate(item.keys(), start=2):
                worksheet.cell(row=row, column=col, value=item[key])

        except Exception as err:
            logger.error(f"{await fanc_name()} {repr(err)}")
            continue

    try:
        workbook.save(doc_path)
        return True

    except Exception as err:
        logger.error(f"{await fanc_name()} {repr(err)}")
        return False


async def create_xlsx(full_act_path: str, doc_path: str):
    """Создание файла xlsx по пути full_act_path
    """
    if not sa_check_path(full_act_path):
        return

    is_created: bool = await create_new_xlsx(report_file=doc_path)
    if is_created is None:
        return

    workbook: Workbook = await get_workbook(fill_report_path=doc_path)
    if workbook is None:
        return

    worksheet: Worksheet = await get_worksheet(workbook, index=0)
    if worksheet is None:
        return

    return workbook, worksheet


async def create_new_xlsx(report_file: str) -> bool:
    """Создание xlsx
    """
    try:
        wb = openpyxl.Workbook()
        wb.save(report_file)
        return True

    except Exception as err:
        logger.error(f"{await fanc_name()} {repr(err)}")
        return False


async def get_workbook(fill_report_path: str):
    """Открыть и загрузить Workbook
    :param fill_report_path: str полный путь к файлу
    :return: Workbook or None
    """
    try:
        workbook: Workbook = openpyxl.load_workbook(fill_report_path)
        return workbook

    except Exception as err:
        logger.error(f"{await fanc_name()} {repr(err)}")
        return None


async def get_worksheet(wb: Workbook, index: int = 0):
    """Получение Страницы из документа по индексу
    :param wb: Workbook - книга xls
    :param index: int - индекс листа
    :return: worksheet or None
    """
    try:
        worksheet: Worksheet = wb.worksheets[index]
        return worksheet

    except Exception as err:
        logger.error(f"{await fanc_name()} {repr(err)}")
        return None


async def write_json(file_path: str, data) -> bool:
    """Запись данных в json

    :param file_path: полный путь для записи / сохранения файла включая расширение,
    :param data: данные для записи / сохранения
    :return: True or False
    """
    try:
        with io.open(file_path, 'w', encoding='utf8') as outfile:
            str_ = json.dumps(data,
                              indent=4,
                              sort_keys=True,
                              separators=(',', ': '),
                              ensure_ascii=False)
            outfile.write(str_)
            return True

    except TypeError as err:
        logger.error(f"TypeError: {repr(err)}")
        return False


async def get_today() -> str:
    return (datetime.today() + timedelta(hours=0)).strftime("%d.%m.%Y")


async def fanc_name():
    stack = traceback.extract_stack()
    return str(stack[-2][2])


async def test():
    user_id: int = 373084462
    test_path = input()
    await process_description(user_id=user_id, test_path=test_path)


if __name__ == "__main__":
    asyncio.run(test())
