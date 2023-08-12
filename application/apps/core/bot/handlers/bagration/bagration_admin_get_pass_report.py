from __future__ import annotations

import sqlite3
import traceback
from datetime import datetime, timedelta, date
import io
import json
import os
import asyncio
from itertools import chain
from openpyxl.styles import Border, Side, Alignment, Font
from xlsxwriter.worksheet import Worksheet

import openpyxl
from aiogram import types

from apps.MyBot import MyBot, bot_send_message, bot_send_document
from apps.core.bot.bot_utils.check_user_registration import check_user_access
# from apps.core.bot.handlers.photo.qr_personal_id_processing import check_ids
from apps.core.bot.keyboards.inline.build_castom_inlinekeyboard import posts_cb
from apps.core.bot.messages.messages import Messages
from apps.core.bot.messages.messages_test import msg
from apps.core.database.DataBase import DataBase
from apps.core.database.query_constructor import QueryConstructor
from apps.core.settyngs import get_sett, BASE_DIR
from apps.core.utils.generate_report.create_xlsx.create_xlsx import create_xlsx
from apps.core.utils.generate_report.create_xlsx.xlsx_config import MAXIMUM_COLUMN_WIDTH, MAXIMUM_ROW_HEIGHT
from config.config import DATA_BASE_BAGRATION_EMPLOYEE_ID
from loader import logger


class DataBaseEmployeeID:

    def __init__(self):

        if not os.path.exists(DATA_BASE_BAGRATION_EMPLOYEE_ID):
            logger.error(f'Path {DATA_BASE_BAGRATION_EMPLOYEE_ID} is not exists!')

        self.db_file: str = DATA_BASE_BAGRATION_EMPLOYEE_ID
        self.connection = sqlite3.connect(self.db_file)
        self.cursor = self.connection.cursor()

    async def get_table_headers(self, table_name: str = None) -> list[str]:
        """Получение всех заголовков таблицы core_violations

        :return: list[ ... ]
        """
        if not table_name:
            return []

        with self.connection:
            result: list = self.cursor.execute(f"PRAGMA table_info('{table_name}')").fetchall()
            clean_headers: list = [item[1] for item in result]
            return clean_headers

    async def get_all_tables_names(self) -> list:
        """Получение всех имен таблиц в БД

        :return:
        """
        with self.connection:
            result: list = self.cursor.execute("SELECT name FROM sqlite_master WHERE type='table';").fetchall()
            return result

    async def get_data_list(self, query: str = None) -> list:
        """Получение данных из таблицы по запросу 'query'"""
        if not query:
            return []

        with self.connection:
            return self.cursor.execute(query).fetchall()

    async def db_get_data_list(self, table_name: str = None, query: str = '') -> list:
        """Получение данных из таблицы table_name"""

        if not query:
            query_kwargs: dict = {
                "action": 'SELECT', "subject": '*',
            }
            query: str = await QueryConstructor(None, table_name, **query_kwargs).prepare_data()

        try:
            with self.connection:
                result: list = self.cursor.execute(query).fetchall()

            if not result:
                # logger.error(f"no matches found {table_name} in DB "
                #              f"because .cursor.execute is none `table_name`: {table_name}")
                return []
            return result
        except (ValueError, sqlite3.OperationalError) as err:
            logger.error(f'Invalid query. {repr(err)}')
            return []
        finally:
            self.cursor.close()


@MyBot.dp.callback_query_handler(posts_cb.filter(action=['bagration_admin_get_pass_report']))
async def call_bagration_admin_get_pass_report(call: types.CallbackQuery = None, callback_data: dict = None,
                                               user_id: int | str = None):
    """Обработка ответов содержащихся в ADMIN_MENU_LIST
    """

    hse_user_id = call.message.chat.id if call else user_id
    if call: logger.debug(f'{hse_user_id = } {call.data = }')

    if not await check_user_access(chat_id=hse_user_id):
        logger.error(f'access fail {hse_user_id = }')
        return

    if not get_sett(cat='enable_features', param='use_bagration_admin_get_pass_report').get_set():
        msg_text: str = f"{await msg(hse_user_id, cat='error', msge='features_disabled', default=Messages.Error.features_disabled).g_mas()}"
        await bot_send_message(chat_id=hse_user_id, text=msg_text, disable_web_page_preview=True)
        return

    # msg_text = await msg(hse_user_id, cat='error', msge='error_action', default=Messages.Error.error_action).g_mas()
    # await bot_send_message(chat_id=hse_user_id, text=msg_text)

    reply_markup = await add_correct_inline_keyboard_with_action()
    msg_text: str = 'Выберите действие'
    await bot_send_message(chat_id=hse_user_id, text=msg_text, reply_markup=reply_markup)


async def add_correct_inline_keyboard_with_action(markup: types.InlineKeyboardMarkup = None):
    """Формирование сообщения с текстом и кнопками действий в зависимости от параметров

    :return:
    """
    if not markup:
        markup = types.InlineKeyboardMarkup()

    markup.add(types.InlineKeyboardButton(
        'За сегодня',
        callback_data=posts_cb.new(id='-', action='use_bagration_admin_get_pass_report_today'))
    )
    markup.add(types.InlineKeyboardButton(
        'За сегодня и вчера',
        callback_data=posts_cb.new(id='-', action='use_bagration_admin_get_pass_report_today_and_yesterday'))
    )
    markup.add(types.InlineKeyboardButton(
        'За текущую неделю',
        callback_data=posts_cb.new(id='-', action='use_bagration_admin_get_pass_report_week'))
    )
    markup.add(types.InlineKeyboardButton(
        'За текущий месяц',
        callback_data=posts_cb.new(id='-', action='use_bagration_admin_get_pass_report_month'))
    )
    markup.add(types.InlineKeyboardButton(
        'За весь период',
        callback_data=posts_cb.new(id='-', action='use_bagration_admin_get_pass_report_all'))
    )
    return markup


@MyBot.dp.callback_query_handler(posts_cb.filter(action=['use_bagration_admin_get_pass_report_today']))
async def call_bagration_admin_get_pass_report(call: types.CallbackQuery = None, callback_data: dict = None,
                                               user_id: int | str = None):
    """Обработка ответов содержащихся в ADMIN_MENU_LIST
    """

    hse_user_id = call.message.chat.id if call else user_id
    if call: logger.debug(f'{hse_user_id = } {call.data = }')

    if not await check_user_access(chat_id=hse_user_id):
        logger.error(f'access fail {hse_user_id = }')
        return

    # msg_text = await msg(hse_user_id, cat='error', msge='error_action', default=Messages.Error.error_action).g_mas()
    # await bot_send_message(chat_id=hse_user_id, text=msg_text)

    now = datetime.now()
    period: list = [now.strftime("%d.%m.%Y"), now.strftime("%d.%m.%Y"), ]
    logger.info(f"{hse_user_id = } {period = }")

    await get_pass_report(hse_user_id, period)


@MyBot.dp.callback_query_handler(posts_cb.filter(action=['use_bagration_admin_get_pass_report_today_and_yesterday']))
async def call_bagration_admin_get_pass_report(call: types.CallbackQuery = None, callback_data: dict = None,
                                               user_id: int | str = None):
    """Обработка ответов содержащихся в ADMIN_MENU_LIST
    """

    hse_user_id = call.message.chat.id if call else user_id
    if call: logger.debug(f'{hse_user_id = } {call.data = }')

    if not await check_user_access(chat_id=hse_user_id):
        logger.error(f'access fail {hse_user_id = }')
        return

    # msg_text = await msg(hse_user_id, cat='error', msge='error_action', default=Messages.Error.error_action).g_mas()
    # await bot_send_message(chat_id=hse_user_id, text=msg_text)

    now = datetime.now()
    previous = now - timedelta(days=1)
    period: list = [previous.strftime("%d.%m.%Y"), now.strftime("%d.%m.%Y"), ]
    logger.info(f"{hse_user_id = } {period = }")

    await get_pass_report(hse_user_id, period)


@MyBot.dp.callback_query_handler(posts_cb.filter(action=['use_bagration_admin_get_pass_report_week']))
async def call_bagration_admin_get_pass_report(call: types.CallbackQuery = None, callback_data: dict = None,
                                               user_id: int | str = None):
    """Обработка ответов содержащихся в ADMIN_MENU_LIST
    """

    hse_user_id = call.message.chat.id if call else user_id
    if call: logger.debug(f'{hse_user_id = } {call.data = }')

    if not await check_user_access(chat_id=hse_user_id):
        logger.error(f'access fail {hse_user_id = }')
        return

    # msg_text = await msg(hse_user_id, cat='error', msge='error_action', default=Messages.Error.error_action).g_mas()
    # await bot_send_message(chat_id=hse_user_id, text=msg_text)

    now = datetime.now()
    current_week: str = await get_week_message(current_date=now)
    current_year: str = await get_year_message(current_date=now)

    period = await db_get_period_for_current_week(current_week, current_year)
    logger.info(f"{hse_user_id = } {period = }")

    await get_pass_report(hse_user_id, period)


@MyBot.dp.callback_query_handler(posts_cb.filter(action=['use_bagration_admin_get_pass_report_month']))
async def call_bagration_admin_get_pass_report(call: types.CallbackQuery = None, callback_data: dict = None,
                                               user_id: int | str = None):
    """Обработка ответов содержащихся в ADMIN_MENU_LIST
    """

    hse_user_id = call.message.chat.id if call else user_id
    if call: logger.debug(f'{hse_user_id = } {call.data = }')

    if not await check_user_access(chat_id=hse_user_id):
        logger.error(f'access fail {hse_user_id = }')
        return

    # msg_text = await msg(hse_user_id, cat='error', msge='error_action', default=Messages.Error.error_action).g_mas()
    # await bot_send_message(chat_id=hse_user_id, text=msg_text)

    now = datetime.now()
    current_month: str = await get_month_message(current_date=now)
    current_year: str = await get_year_message(current_date=now)

    period = await db_get_period_for_current_month(current_month, current_year)
    logger.info(f"{hse_user_id = } {period = }")

    await get_pass_report(hse_user_id, period)


@MyBot.dp.callback_query_handler(posts_cb.filter(action=['use_bagration_admin_get_pass_report_all']))
async def call_bagration_admin_get_pass_report(call: types.CallbackQuery = None, callback_data: dict = None,
                                               user_id: int | str = None):
    """Обработка ответов содержащихся в ADMIN_MENU_LIST
    """

    hse_user_id = call.message.chat.id if call else user_id
    if call: logger.debug(f'{hse_user_id = } {call.data = }')

    if not await check_user_access(chat_id=hse_user_id):
        logger.error(f'access fail {hse_user_id = }')
        return

    # msg_text = await msg(hse_user_id, cat='error', msge='error_action', default=Messages.Error.error_action).g_mas()
    # await bot_send_message(chat_id=hse_user_id, text=msg_text)

    period = None

    await get_pass_report(hse_user_id, period)


async def get_pass_report(hse_user_id, period: list = None):
    """"""
    media_patch: str = f'{BASE_DIR.parent.parent}\\media\\BAGRATION\\personal_id_hunting\\'

    if not os.path.isdir(media_patch):
        os.makedirs(media_patch)

    json_files_list: list = []
    for subdir, dirs, files in os.walk(media_patch):
        for file in files:

            filepath = subdir + os.sep + file
            if filepath.endswith('.py'): continue
            if filepath.endswith('.jpg'): continue
            if filepath.endswith('.tmp'): continue
            if filepath.endswith('.db'): continue
            if "personal_id_data.json" not in filepath: continue
            if not file: continue
            if '~' in file: continue

            if not await check_data_period(file, period): continue

            json_files_list.append(
                {
                    "file": file,
                    "subdir": subdir,
                    "full_file_path": f'{subdir}\\{file}',
                }
            )
    all_data_id_list: list = [
        await read_json_file(item.get('full_file_path')) for item in json_files_list if item.get('full_file_path')
    ]

    id_data_list: list = list(chain(*[list(item) for item in all_data_id_list]))

    employee_list: list = []
    for i in id_data_list:
        employee_list.append([k for k, v in i.items()][0])

    uniq_employee_list_fin: list = list(set(employee_list))

    msg_text: str = f'Количество записей за выбранный период: {len(uniq_employee_list_fin)}'
    await bot_send_message(chat_id=hse_user_id, text=msg_text)

    if not len(uniq_employee_list_fin):
        return

    main_path: str = f'{media_patch}!reports\\'
    if not os.path.isdir(main_path):
        os.makedirs(main_path)

    current_date: str = datetime.now().strftime("%d.%m.%Y")
    full_doc_path: str = await create_personal_id_xlsx(
        hse_user_id, uniq_employee_list_fin, main_path, deta=current_date, period=period
    )

    if not period: period: list = ['все записи']
    period_text = ' - '.join(period)

    caption: str = f"Сформирован отчет за период: {period_text} "
    kwargs: dict = {
        'reply_markup': await add_employee_inline_keyboard_with_action()
    }

    await bot_send_document(
        chat_id=hse_user_id,
        doc_path=full_doc_path,
        caption=caption,
        fanc_name=await fanc_name(),
        **kwargs
    )

    return uniq_employee_list_fin


async def create_personal_id_xlsx(chat_id, data_list, main_path, deta, period):
    """"""

    if not period:
        period = []

    report_full_path: str = f"{main_path}Отчет от {deta} personal_id_data за период {' - '.join(period)}.xlsx"

    workbook, worksheet = await create_xlsx(chat_id, report_full_path)
    if not worksheet:
        return False

    await format_sheets(worksheet)
    row_num = 0
    for row, item in enumerate(data_list, start=2):

        id_dict = await check_ids(item)

        if not id_dict.get('emploee_id'):
            continue

        row_num = + 1
        try:
            worksheet.cell(row=row_num, column=1, value=row_num)
            worksheet.cell(row=row_num, column=2, value=id_dict.get('emploee_id'))
            worksheet.cell(row=row_num, column=3, value=id_dict.get('subcontractor'))
            worksheet.cell(row=row_num, column=4, value=id_dict.get('function'))

        except Exception as err:
            logger.error(f"set_user_values {repr(err)}")
            continue

    workbook.save(report_full_path)
    return report_full_path


async def check_ids(item_id) -> dict:
    """"""
    list_datas = await test_fank(item_id)
    id_data_list: list = list(chain(*[list(item) for item in list_datas]))
    fjgh = list(set([item.get('id') for item in id_data_list]))
    if not fjgh:
        return {}
    id_data: list = await check_employee_id_from_db(fjgh[0])
    return id_data[0]


async def test_fank(item_id):
    list_1_da = 0
    list_1_net = 0

    list_3_da = 0
    list_3_net = 0
    list_1_da_List = []

    emploee_id = item_id[1:-2]
    item_dict = await check_employee_id(emploee_id)
    # print(f'{len(item_dict)} item_id[1:-2] {emploee_id} {item_dict} ')
    if len(item_dict) == 1:
        list_1_da += 1
        # print(item_dict)
        list_1_da_List.append(item_dict)

    if len(item_dict) > 1:
        list_1_net += 1

    emploee_id = item_id[1:-1]
    item_dict = await check_employee_id(emploee_id)
    # print(f'{len(item_dict)} item_id[1:-1] {emploee_id}  {item_dict}')
    if len(item_dict) == 1:
        list_3_da += 1
        # print(item_dict)
        list_1_da_List.append(item_dict)

    if len(item_dict) > 1:
        list_3_net += 1

    # print(f'{list_1_da = }')
    # print(f'{list_1_net = }')
    # print(f'{list_3_da = }')
    # print(f'{list_3_net = }')
    #
    # print(f'{len(list_1_da_List)}')
    return list_1_da_List


async def check_employee_id(item_id) -> list:
    """"""
    table_name = 'bagration_subcontractor_emploee_id'

    if not item_id:
        return []

    query_kwargs: dict = {
        "action": 'SELECT', "subject": '*',
        "conditions": {
            "lazy_query": f"`emploee_id` LIKE '%{item_id}%'",
        },
    }
    query: str = await QueryConstructor(None, table_name, **query_kwargs).prepare_data()

    try:
        datas_list: list = await DataBaseEmployeeID().db_get_data_list(table_name=table_name, query=query)

    except sqlite3.OperationalError as err:
        # message = ''
        # logger.error(f'The "{message}" {table_name = } is missing from the database! err: {repr(err)}')
        return []

    if not datas_list:
        # logger.error('Missing datas_list from the database.')
        return []

    clean_headers: list = await DataBaseEmployeeID().get_table_headers(table_name)
    list_dicts: list = [dict(zip(clean_headers, data_list)) for data_list in datas_list]

    if list_dicts:
        return list_dicts

    return []


async def check_employee_id_from_db(item_id) -> list:
    """"""
    table_name = 'bagration_subcontractor_emploee_id'

    if not item_id:
        return []

    query_kwargs: dict = {
        "action": 'SELECT', "subject": '*',
        "conditions": {
            "id": item_id,
        },
    }
    query: str = await QueryConstructor(None, table_name, **query_kwargs).prepare_data()

    try:
        datas_list: list = await DataBaseEmployeeID().db_get_data_list(table_name=table_name, query=query)

    except sqlite3.OperationalError as err:
        # message = ''
        # logger.error(f'The "{message}" {table_name = } is missing from the database! err: {repr(err)}')
        return []

    if not datas_list:
        # logger.error('Missing datas_list from the database.')
        return []

    clean_headers: list = await DataBaseEmployeeID().get_table_headers(table_name)
    list_dicts: list = [dict(zip(clean_headers, data_list)) for data_list in datas_list]

    if list_dicts:
        return list_dicts

    return []


async def add_employee_inline_keyboard_with_action():
    """Формирование сообщения с текстом и кнопками действий в зависимости от параметров

    :return:
    """

    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton(
        text='не нажимать',
        callback_data=posts_cb.new(id='-', action='')))

    return markup


async def check_data_period(file: str, period: list = None) -> bool:
    """"""
    if not period: return True
    if len(period) == 1: return False
    if len(period) > 2: return False

    _format: str = '%d.%m.%Y'
    file_date_from_name = file.split(' ')[0]

    try:
        file_date = datetime.strptime(file_date_from_name, _format).date()
    except ValueError:
        logger.warning(f"file: {file} file_date: {file.split(' ')[0]}")
        return False

    start_period = datetime.strptime(period[0], _format).date()
    stop_period = datetime.strptime(period[1], _format).date()

    if start_period <= file_date <= stop_period:
        return True

    return False


async def read_json_file(file: str):
    """Получение данных из json.

    :param file: полный путь к файлу
    """
    try:
        with open(file, 'r', encoding='utf8') as data_file:
            data_loaded = json.load(data_file)
        return data_loaded
    except FileNotFoundError:
        return None


async def write_json(name: str, data: list) -> bool:
    """Запись данных в json

    :param name: полный путь для записи / сохранения файла включая расширение,
    :param data: данные для записи / сохранения
    :return: True or False
    """
    try:
        with io.open(name, 'w', encoding='utf8') as outfile:
            str_ = json.dumps(data,
                              indent=4,
                              sort_keys=True,
                              separators=(',', ': '),
                              ensure_ascii=False)
            outfile.write(str_)
            return True
    except TypeError as err:
        logger.error(f"TypeError: {repr(err)}")
        return False


async def get_week_message(current_date: datetime | str = None) -> str:
    """Обработчик сообщений с фото
    Получение номер str недели из сообщения в формате dd
    """
    current_date: date = await str_to_datetime(current_date)

    if not current_date:
        current_date: datetime = datetime.now()
    week = current_date.isocalendar()[1]
    return str("0" + str(week) if week < 10 else str(week))


async def get_month_message(current_date: datetime = None) -> str:
    """Получение номер str месяца из сообщения в формате mm
    """
    current_date: date = await str_to_datetime(current_date)

    if not current_date:
        current_date: datetime = datetime.now()
    return str("0" + str(current_date.month) if int(current_date.month) < 10 else str(current_date.month))


async def get_year_message(current_date: datetime = None) -> str:
    """Обработчик сообщений с фото
    Получение полного пути файла
    """
    current_date: date = await str_to_datetime(current_date)

    if not current_date:
        current_date: datetime = datetime.now()

    return str(current_date.year)


async def str_to_datetime(date_str: str | datetime) -> date:
    """Преобразование str даты в datetime

    :param
    """

    current_date: date = None
    try:
        if isinstance(date_str, str):
            current_date: date = datetime.strptime(date_str, "%d.%m.%Y").date()
    except ValueError as err:
        logger.error(f"{repr(err)}")

    return current_date


async def db_get_period_for_current_week(current_week: str, current_year: str = None) -> list:
    """Получение данных из core_week по week_number

    :return:
    """
    if not current_week:
        logger.error('ERROR: No user_id foe db_get_username ')

    if not current_year:
        current_year = await get_year_message(current_date=datetime.now())

    query_kwargs: dict = {
        "action": 'SELECT', "subject": '*',
        "conditions": {
            "week_number": current_week,
        },
    }
    query: str = await QueryConstructor(None, 'core_week', **query_kwargs).prepare_data()

    logger.debug(f'{__name__} {await fanc_name()} {query}')

    datas_query: list = DataBase().get_data_list(query=query)
    period_data = datas_query[0]

    table_headers: list = DataBase().get_table_headers('core_week')
    headers = [row[1] for row in table_headers]
    period_dict = dict(zip(headers, period_data))
    return [
        period_dict.get(f'start_{current_year}', None),
        period_dict.get(f'end_{current_year}', None)
    ]


async def db_get_period_for_current_month(current_month: str = None, current_year: str = None) -> list:
    """Получение данных из core_week по week_number

    :return:
    """

    if not current_month:
        current_month = await get_month_message(current_date=datetime.now())

    if not current_year:
        current_year = await get_year_message(current_date=datetime.now())

    last_day_month = last_day_of_month(current_date=datetime.now())
    last_day_month = last_day_month.strftime("%d.%m.%Y")

    period = [
        f'01.{current_month}.{current_year}',
        f'{last_day_month}',
    ]
    return period


def last_day_of_month(current_date: datetime):
    if current_date.month == 12:
        return current_date.replace(day=31)
    return current_date.replace(month=current_date.month + 1, day=1) - timedelta(days=1)


async def fanc_name():
    stack = traceback.extract_stack()
    return str(stack[-2][2])


async def format_sheets(worksheet: Worksheet):
    """Пошаговое форматирование страницы
    """

    await set_border(worksheet)
    await set_alignment(worksheet)
    await set_font(worksheet)
    await set_column_widths(worksheet)
    await set_row_height(worksheet)


async def set_border(worksheet):
    """Форматирование ячейки: все границы ячейки

    """
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for row in worksheet.iter_rows():
        for cell in row:
            try:
                cell.border = thin_border
            except Exception as err:
                logger.error(f"set_border {repr(err)}")


async def set_alignment(worksheet):
    """Форматирование ячейки: положение текста в ячейке (лево верх)
    """
    wrap_alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

    for row in worksheet.iter_rows():
        for cell in row:
            try:
                cell.alignment = wrap_alignment
            except Exception as err:
                logger.error(f"set_alignment {repr(err)}")


async def set_font(worksheet) -> bool:
    """Форматирование ячейки: размер шрифта

    """
    for row in worksheet.iter_rows():
        for cell in row:
            try:
                cell.font = Font(size=14)
            except Exception as err:
                logger.error(f"sets_report_font {repr(err)}")
                continue
    return True


async def set_column_widths(worksheet):
    """Форматирование ячейки: ширина столбца

    """

    for column_cells in worksheet.columns:
        # максимальная ширина столбца
        column_length = max(len(_as_text(cell.value)) for cell in column_cells)

        if column_length < MAXIMUM_COLUMN_WIDTH:
            new_column_length = column_length
        else:
            new_column_length = MAXIMUM_COLUMN_WIDTH

        new_column_letter: int = (openpyxl.utils.get_column_letter(column_cells[0].column))
        if new_column_length > 0:
            try:
                worksheet.column_dimensions[new_column_letter].width = new_column_length + 1
            except Exception as err:
                logger.error(f"set_column_widths {repr(err)}")


def _as_text(value) -> str:
    """Приведение данных к str

    """
    if value is None:
        return ""
    return str(value)


async def set_row_height(worksheet):
    """Форматирование ячейки: высота шрифта
    """
    for ind in range(worksheet.max_row):
        if ind == 0:
            continue
        try:
            worksheet.row_dimensions[ind + 1].height = MAXIMUM_ROW_HEIGHT
        except Exception as err:
            logger.error(F"set_row_height {repr(err)}")


async def test():
    # await call_bagration_admin_get_pass_report()

    hse_user_id = 373084462
    period = None  # [datetime.now().strftime("%d.%m.%Y"), datetime.now().strftime("%d.%m.%Y")]

    await get_pass_report(hse_user_id, period)

    return []


if __name__ == '__main__':
    asyncio.run(test())
