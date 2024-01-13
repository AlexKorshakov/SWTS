from __future__ import annotations

import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.worksheet import Worksheet

from pandas import DataFrame

from apps.MyBot import bot_send_message
from apps.core.bot.messages.messages import Messages
from apps.core.bot.reports.report_data import headlines_data
from apps.core.database.db_utils import (db_get_data_dict_from_table_with_id,
                                         db_get_data_list,
                                         db_get_dict_userdata,
                                         db_get_categories_list)
from loader import logger

check_mark_true: str = 'V'
not_found: str = 'не выявлено'
not_tested: str = 'не проверялось'


async def create_xlsx(chat_id: int, full_act_path: str):
    """

    :param chat_id:
    :param full_act_path:
    :return:
    """

    is_created: bool = await create_new_xlsx(report_file=full_act_path)
    if is_created is None:
        logger.warning(Messages.Error.workbook_not_create)
        await bot_send_message(chat_id=chat_id, text=Messages.Error.workbook_not_create)
        return

    workbook = await get_workbook(fill_report_path=full_act_path)
    if workbook is None:
        logger.warning(Messages.Error.workbook_not_found)
        await bot_send_message(chat_id=chat_id, text=Messages.Error.workbook_not_found)
        return

    worksheet = await get_worksheet(workbook, index=0)
    if worksheet is None:
        logger.warning(Messages.Error.worksheet_not_found)
        await bot_send_message(chat_id=chat_id, text=Messages.Error.worksheet_not_found)
        return

    return workbook, worksheet


async def create_new_xlsx(report_file: str) -> bool:
    """Создание xlsx
    """
    try:
        wb = openpyxl.Workbook()
        wb.save(report_file)
        return True

    except Exception as err:
        logger.error(F"set_border {repr(err)}")
        return False


async def get_worksheet(wb: Workbook, index: int = 0) -> Worksheet | None:
    """Получение Страницы из документа по индексу
    :param wb: Workbook - книга xls
    :param index: int - индекс листа
    :return: worksheet or None
    """
    try:
        worksheet: Worksheet = wb.worksheets[index]
        return worksheet

    except Exception as err:
        logger.error(f"get_workbook {repr(err)}")
        return None


async def get_workbook(fill_report_path: str) -> Workbook | None:
    """Открыть и загрузить Workbook
    :param fill_report_path: str полный путь к файлу
    :return: Workbook or None
    """
    try:
        workbook: Workbook = openpyxl.load_workbook(fill_report_path)
        return workbook

    except Exception as err:
        logger.error(f"get_workbook {repr(err)}")
        return None


async def format_stat_sheet_basic(worksheet: Worksheet) -> bool:
    """

    :param worksheet:
    :return:
    """
    await set_page_setup(worksheet)

    for item in STATISTIC_RANGE_COLUMNS_BASIC:
        await set_column_dimensions(worksheet, column=item[0], width=item[1])

    for item_range in STATISTIC_CELL_RANGES_BASIC_ALIGNMENT:
        await set_report_font(worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1])

    for cell_range in STATISTIC_CELL_RANGES_BASIC_ALIGNMENT:
        await set_stat_alignment(worksheet, cell_range[0], horizontal='left', vertical='center')

    return True


async def set_stat_values_header(worksheet: Worksheet) -> bool:
    """Заполнение первоначальных данных отчета

    :param worksheet: страница для заполнения
    :return:
    """
    values = [
        {"coordinate": "C2", "value": "НАЗВАНИЕ ОРГАНИЗАЦИИ", "row": "2", "column": "3"},
        # {"coordinate": "D2", "value": "Отчет о ночной смены", "row": "2", "column": "4"},
        {"coordinate": "C3", "value": "СТАТ-ОТПБиПТ--", "row": "3", "column": "3"},
        {"coordinate": "D3", "value": "Значение", "row": "3", "column": "4"},
        {"coordinate": "F3", "value": "Примечание", "row": "3", "column": "6"},
        {"coordinate": "C4", "value": "Общая информация", "row": "4", "column": "3"},
        {"coordinate": "C5", "value": "Отчет", "row": "5", "column": "3"},
        {"coordinate": "D5", "value": "Период: ", "row": "5", "column": "4"},
        {"coordinate": "C6", "value": "Дата составления", "row": "6", "column": "3"},
        {"coordinate": "C7", "value": "Подрядчик(и)", "row": "7", "column": "3"},
        # {"coordinate": "D7", "value": "стм.", "row": "7", "column": "4"},
        # {"coordinate": "C8", "value": "Субподрядчик", "row": "8", "column": "3"},
        {"coordinate": "C9", "value": "Проект", "row": "9", "column": "3"},
        # {"coordinate": "D9", "value": "", "row": "9", "column": "4"},
        {"coordinate": "C10", "value": "Комиссия", "row": "10", "column": "3"},
        {"coordinate": "D10", "value": "Функция", "row": "10", "column": "4"},
        {"coordinate": "F10", "value": "ФИО", "row": "10", "column": "6"},
        {"coordinate": "D11", "value": "□", "row": "11", "column": "4"},
        # {"coordinate": "E11", "value": "Инспектирующие", "row": "11", "column": "5"},
        {"coordinate": "D12", "value": "□", "row": "12", "column": "4"},
        # {"coordinate": "E12", "value": "Руководитель строительства", "row": "12", "column": "5"},
        # {"coordinate": "F12", "value": f"{be_away}", "row": "12", "column": "6"},
        {"coordinate": "D13", "value": f"{check_mark_true}", "row": "13", "column": "4"},
        {"coordinate": "E13", "value": "Специалист отдела ПБ", "row": "13", "column": "5"},
        {"coordinate": "D14", "value": "□", "row": "14", "column": "4"},
        # {"coordinate": "E14", "value": "Инженер СК", "row": "14", "column": "5"},
        # {"coordinate": "F14", "value": f"{be_away}", "row": "14", "column": "6"},
        {"coordinate": "D15", "value": "□", "row": "15", "column": "4"},
        # {"coordinate": "E15", "value": "Подрядчик", "row": "15", "column": "5"},
        # {"coordinate": "F15", "value": f"{be_away}", "row": "15", "column": "6"},
        {"coordinate": "D16", "value": "□", "row": "16", "column": "4"},
        # {"coordinate": "E16", "value": "Субподрядчик", "row": "16", "column": "5"},
        {"coordinate": "C17", "value": "Охрана труда, промышленная безопасность и охрана окружающей среды", "row": "17",
         "column": "3"},
        {"coordinate": "C18", "value": "Наблюдения", "row": "18", "column": "3"},
        {"coordinate": "D18", "value": "Категория несоответствия", "row": "18", "column": "4"},
        {"coordinate": "F18", "value": "№", "row": "18", "column": "6"},
        {"coordinate": "G18", "value": "Количество (всего / в актах )", "row": "18", "column": "7"},
        {"coordinate": "H18", "value": "Не устраненных / просрочено", "row": "18", "column": "8"},
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column'])).value = str(val['value'])

            if val["value"] == not_found:
                cell_range = [f"D{val['row']}:H{val['row']}",
                              {"color": "008000", "font_size": "12", "name": "Arial"}]
                await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

        except Exception as err:
            logger.error(f"set_values {repr(err)}")
            return False
    return True


async def format_stat_sheet_header(worksheet: Worksheet) -> bool:
    """Пошаговое форматирование страницы
    """

    for merged_cell in STATISTIC_MERGED_CELLS_HEADER:
        await set_merge_cells(worksheet, merged_cell=merged_cell)

    for item in STATISTIC_CELL_RANGES_HEADER:
        await set_range_border(worksheet, cell_range=item[0], border=item[1])

    for item in STATISTIC_ROW_DIMENSIONS_HEADER:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    for item in STATISTIC_CELL_RANGE_BACKGROUND_COLOR_HEADER:
        await set_report_background_color(worksheet, item[0], rgb=item[1])

    for cell_range in STATISTIC_CELL_RANGES_ALIGNMENT_HEADER:
        await set_stat_alignment(worksheet, cell_range[0], horizontal='center', vertical='center')

    for cell_range in STATISTIC_CELL_RANGES_SET_REPORT_FONT_HEADER:
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

    return True


async def set_stat_values_body(worksheet: Worksheet) -> tuple[bool, int, list]:
    """Заполнение первоначальных данных отчета

    :param worksheet: страница для заполнения
    :return:
    """
    values: list = []
    num: int = 0

    # TODO заменить на вызов конструктора QueryConstructor
    query: str = "SELECT `title` FROM `core_category`"
    categories: list = await db_get_data_list(query=query)
    clean_title_categories = [item[0] for item in categories]

    for num, category in enumerate(clean_title_categories, start=19):
        values.append({"coordinate": f"D{num}", "value": f"{check_mark_true}", "row": f"{num}", "column": "4"})
        values.append({"coordinate": f"E{num}", "value": f"{category}", "row": f"{num}", "column": "5"})
        values.append({"coordinate": f"G{num}", "value": f"{not_found}", "row": f"{num}", "column": "7"})

    values.append({"coordinate": f"D{num + 1}", "value": f"{check_mark_true}", "row": f"{num + 1}", "column": "4"})
    values.append({"coordinate": f"E{num + 1}", "value": "В С Е Г О", "row": f"{num + 1}", "column": "5"})
    values.append({"coordinate": f"G{num + 1}", "value": f"{not_found}", "row": f"{num + 1}", "column": "7"})

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column'])).value = str(val['value'])

            if val["value"] == not_found:
                cell_range = [f"D{val['row']}:H{val['row']}",
                              {"color": "008000", "font_size": "14", "name": "Arial"}]
                await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

        except Exception as err:
            logger.error(f"set_values {repr(err)}")
            return False, num + 1, values

    # TODO заменить на вызов конструктора QueryConstructor
    query: str = "SELECT `title` FROM `core_maincategory`"
    main_categories: list = await db_get_data_list(query=query)
    clean_title_main_categories = [item[0] for item in main_categories]

    m_num = num + 3
    for m_num, m_category in enumerate(clean_title_main_categories, start=num + 3):
        values.append({"coordinate": f"D{m_num}", "value": f"{check_mark_true}", "row": f"{m_num}", "column": "4"})
        values.append({"coordinate": f"E{m_num}", "value": f"{m_category}", "row": f"{m_num}", "column": "5"})
        values.append({"coordinate": f"G{m_num}", "value": f"{not_found}", "row": f"{m_num}", "column": "7"})

    values.append({"coordinate": f"D{m_num + 1}", "value": f"{check_mark_true}", "row": f"{m_num + 1}", "column": "4"})
    values.append({"coordinate": f"E{m_num + 1}", "value": "*В С Е Г О*", "row": f"{m_num + 1}", "column": "5"})
    values.append({"coordinate": f"G{m_num + 1}", "value": f"{not_found}", "row": f"{m_num + 1}", "column": "7"})

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column'])).value = str(val['value'])

            if val["value"] == not_found:
                cell_range = [f"D{val['row']}:H{val['row']}",
                              {"color": "008000", "font_size": "14", "name": "Arial"}]
                await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

        except Exception as err:
            logger.error(f"set_values {repr(err)}")
            return False, m_num + 1, values
    return True, m_num + 1, values


async def format_stat_sheet_body(worksheet: Worksheet, workbook, full_stat_path,
                                 num: int) -> bool:
    """Пошаговое форматирование страницы

    """

    STATISTIC_MERGED_CELLS_BODY = [
        f'C19:C{num}',
    ]

    for merged_cell in STATISTIC_MERGED_CELLS_BODY:
        await set_merge_cells(worksheet, merged_cell=merged_cell)
    workbook.save(full_stat_path)

    STATISTIC_CELL_RANGES_BODY = [
        (f'C19:H{num}', False),
    ]

    for item in STATISTIC_CELL_RANGES_BODY:
        await set_range_border(worksheet, cell_range=item[0], border=item[1])
    workbook.save(full_stat_path)

    STATISTIC_ROW_DIMENSIONS_BODY: list = []
    for item in range(19, num + 1):
        STATISTIC_ROW_DIMENSIONS_BODY.append([f"{item}", "33.0"])

    for item in STATISTIC_ROW_DIMENSIONS_BODY:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])
    workbook.save(full_stat_path)

    STATISTIC_CELL_RANGE_BACKGROUND_COLOR_BODY = [
        [f"C19:C{num}", "FFDCDAFA"],
    ]
    for item in STATISTIC_CELL_RANGE_BACKGROUND_COLOR_BODY:
        await set_report_background_color(worksheet, item[0], rgb=item[1])
    workbook.save(full_stat_path)

    # STATISTIC_CELL_RANGES_ALIGNMENT_BODY = [
    # ]

    for cell_range in STATISTIC_CELL_RANGES_ALIGNMENT_BODY:
        await set_stat_alignment(worksheet, cell_range[0], horizontal='center', vertical='center')
    workbook.save(full_stat_path)

    # STATISTIC_CELL_RANGES_SET_REPORT_FONT_BODY = [
    # ]

    for cell_range in STATISTIC_CELL_RANGES_SET_REPORT_FONT_BODY:
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])
    workbook.save(full_stat_path)

    return True


async def set_stat_values_footer(worksheet: Worksheet, workbook,
                                 full_daily_report_report_path, num) -> tuple[bool, int]:
    """Заполнение первоначальных данных отчета

    :param num:
    :param full_daily_report_report_path:
    :param workbook:
    :param worksheet: страница для заполнения
    :return:
    """

    values = [
        {"coordinate": f"C{num + 1}", "value": "Дополнительная информация", "row": f"{num + 1}", "column": "3"},
        {"coordinate": f"C{num + 2}",
         "value": 'Данное сообщение рассылается Дирекцией по охране труда, промышленной безопасности и охране '
                  'окружающей среды  с целью информирования о состоянии площадки, производства и '
                  'документирования строительно-монтажных работ',
         "row": f"{num + 2}", "column": "3"}
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column'])).value = str(val['value'])

            if val["value"] == not_found:
                cell_range = [f"D{val['row']}:H{val['row']}",
                              {"color": "008000", "font_size": "14", "name": "Arial"}]
                await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

        except Exception as err:
            logger.error(f"set_values {repr(err)}")
            return False, num

    workbook.save(full_daily_report_report_path)
    return True, num


async def format_stat_sheet_footer(worksheet: Worksheet, workbook,
                                   full_stat_path, num) -> tuple[bool, int]:
    """Пошаговое форматирование страницы
    """
    STATISTIC_MERGED_CELLS_FOOTER = [
        f'C{num + 1}:H{num + 1}',
        f'C{num + 2}:H{num + 2}',
    ]

    for merged_cell in STATISTIC_MERGED_CELLS_FOOTER:
        await set_merge_cells(worksheet, merged_cell=merged_cell)
    workbook.save(full_stat_path)

    STATISTIC_CELL_RANGES_FOOTER = [
        (f'C2:H{num + 1}', False),
        (f'C2:H{num + 2}', False),
    ]

    for item in STATISTIC_CELL_RANGES_FOOTER:
        await set_range_border(worksheet, cell_range=item[0], border=item[1])
    workbook.save(full_stat_path)

    STATISTIC_ROW_DIMENSIONS_FOOTER = [
        [f"{num + 1}", "17.4"],
        [f"{num + 2}", "58.2"],
        [f"{num + 3}", "33.0"],
    ]

    for item in STATISTIC_ROW_DIMENSIONS_FOOTER:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])
    workbook.save(full_stat_path)

    STATISTIC_CELL_RANGE_BACKGROUND_COLOR_FOOTER = [
        [f"C{num + 1}:H{num + 1}", "FFFFC000"],
    ]

    for item in STATISTIC_CELL_RANGE_BACKGROUND_COLOR_FOOTER:
        await set_report_background_color(worksheet, item[0], rgb=item[1])
    workbook.save(full_stat_path)

    STATISTIC_CELL_RANGES_ALIGNMENT_FOOTER = [
        (f"C{num + 1}:H{num + 1}", 'Times New Roman', 14, 'center', 'center'),
        (f"C{num + 2}:H{num + 2}", 'Times New Roman', 14, 'center', 'center'),
    ]

    for cell_range in STATISTIC_CELL_RANGES_ALIGNMENT_FOOTER:
        await set_stat_alignment(worksheet, cell_range[0], horizontal='center', vertical='center')
    workbook.save(full_stat_path)

    STATISTIC_CELL_RANGES_SET_REPORT_FONT_FOOTER = [
        [f"C{num + 1}:H{num + 1}", {"font_size": "14", "bold": "True", "name": "Arial"}],
        [f"C{num + 2}:H{num + 2}", {"font_size": "14", "bold": "True", "name": "Arial"}],
    ]

    for cell_range in STATISTIC_CELL_RANGES_SET_REPORT_FONT_FOOTER:
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])
    workbook.save(full_stat_path)

    return True, num + 2


async def set_row_breaks(worksheet, row: int):
    """Установка горизонтального разрыва страницы

    :return:
    """
    worksheet.row_breaks.append(Break(id=row))


async def set_stat_headlines_data_values(chat_id: int, dataframe=None, query_period=None) -> bool:
    """Формирование заголовков отчета

    :param query_period:
    :param dataframe:
    :param chat_id: id пользователя для которого заполняется отчет
    :return:
    """

    be_away: str = 'отсутствовал'

    if not headlines_data:
        # Руководитель строительства
        headlines_data['construction_manager'] = be_away
        # Инженер СК
        headlines_data['building_control_engineer'] = be_away
        # Подрядчик
        headlines_data['general_contractor'] = ''
        # Подрядчики
        headlines_data['contractors'] = ''
        # Субподрядчик
        headlines_data['subcontractor'] = ''
        # Проект
        headlines_data['name_location'] = ''
        # Вид обхода
        headlines_data['linear_bypass'] = 'за период: ' + ' - '.join(query_period)
        # Дата
        headlines_data['date_linear_bypass'] = ''
        # Представитель подрядчика
        headlines_data['contractor_representative'] = be_away
        # Представитель субподрядчика
        headlines_data['subcontractor_representative'] = be_away

    headlines_data["day"] = (datetime.datetime.now()).strftime("%d")
    headlines_data["year"] = (datetime.datetime.now()).strftime("%Y")

    registration_data: dict = await db_get_dict_userdata(chat_id)

    organization_full_name: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_generalcontractor',
        post_id=registration_data.get("hse_organization", ''))
    organization_name = organization_full_name.get('title', '')

    headlines_data['organization'] = organization_name
    headlines_data['function'] = registration_data.get("hse_function", '')
    headlines_data['name'] = registration_data.get("hse_short_name", '')

    location_list = []
    if not dataframe.empty:

        location = dataframe.location_id.tolist()
        location = list(set(list(location)))

        for item in location:
            if isinstance(item, float):
                continue

            location_full_name: dict = await db_get_data_dict_from_table_with_id(
                table_name='core_location',
                post_id=item)
            location_list.append(location_full_name.get('title', ''))

        headlines_data['name_location'] = ', '.join(location_list)

    if not location_list:
        location_dict: dict = await db_get_data_dict_from_table_with_id(
            table_name='core_location',
            post_id=registration_data.get("hse_location", '')
        )
        location = location_dict.get('title', '')
        headlines_data['name_location'] = location

    headlines_data['phone_number'] = registration_data.get("hse_contact_main_phone_number", '')

    work_shift_dict: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_workshift',
        post_id=registration_data.get("hse_work_shift", '')
    )
    work_shift = work_shift_dict.get('title', '')

    headlines_data['work_shift'] = work_shift

    date_now = datetime.datetime.now().strftime("%d.%m.%Y")
    date_then = datetime.datetime.now() - datetime.timedelta(days=1)
    date_then = date_then.strftime("%d.%m.%Y")

    if headlines_data.get("work_shift", '').lower() == "дневная смена":
        headlines_data['work_shift'] = 'дневной смены'
        headlines_data['custom_date'] = f"{date_now}"
    else:
        headlines_data['work_shift'] = 'о ночной смене'
        headlines_data['custom_date'] = f"{date_then} - {date_now}"

    return True


async def set_stat_header_values(worksheet: Worksheet, dataframe: DataFrame) -> bool:
    """Заполнение заголовка отчета

    :param dataframe:
    :param worksheet: страница для заполнения
    :return:
    """

    contractors = dataframe.general_contractor_id.tolist()
    contractors = list(set(list(contractors)))
    contractors_str = ''
    for item in contractors:
        if isinstance(item, float):
            continue

        general_contractor_full_name: dict = await db_get_data_dict_from_table_with_id(
            table_name='core_generalcontractor',
            post_id=item)
        general_contractor_name = general_contractor_full_name.get('title', '')

        headlines_data['general_contractor_full_name'] = general_contractor_name
        gc_full_name: str = general_contractor_full_name['title']

        if gc_full_name and gc_full_name != 'Подрядная организация':
            contractors_str = contractors_str + gc_full_name + ', '
            headlines_data['contractors'] = contractors_str

    values = [
        {"coordinate": "C2", "value": f"{headlines_data.get('organization')}",
         "row": "2", "column": "3"},
        {"coordinate": "D2",
         "value": f"Статистический отчет {headlines_data.get('work_shift')} {headlines_data.get('custom_date')}",
         "row": "2", "column": "4"},
        {"coordinate": "C3", "value": f"СТАТ-ОТПБиООС-{headlines_data.get('year')}-{headlines_data.get('day')}",
         "row": "3", "column": "3"},
        {"coordinate": "D5", "value": f"{headlines_data.get('linear_bypass')}", "row": "5", "column": "4"},
        {"coordinate": "D6", "value": f"{headlines_data.get('custom_date')}", "row": "6", "column": "4"},

        {"coordinate": "D7", "value": f"{headlines_data.get('contractors')}", "row": "7", "column": "4"},

        {"coordinate": "D8", "value": "", "row": "8", "column": "4"},
        {"coordinate": "D9", "value": f"{headlines_data.get('name_location')}", "row": "9", "column": "4"},
        # {"coordinate": "F12", "value": f"{headlines_data.get('construction_manager')}", "row": "12", "column": "6"},

        {"coordinate": "E13", "value": f"{headlines_data.get('function')}", "row": "13", "column": "5"},
        {"coordinate": "E13", "value": f"{headlines_data.get('name')} тел. {headlines_data.get('phone_number')}",
         "row": "13", "column": "6"},

        # {"coordinate": "F14", "value": f"{headlines_data.get('building_control_engineer')}", "row": "14",
        #  "column": "6"},
        # {"coordinate": "F15", "value": f"{headlines_data.get('contractor_representative')}", "row": "15",
        #  "column": "6"},
        # {"coordinate": "F16", "value": f"{headlines_data.get('subcontractor_representative')}", "row": "16",
        #  "column": "6"},
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column']), value=str(val['value']))

        except Exception as err:
            logger.error(f"set_user_values {repr(err)}")
            return False
    return True


async def set_stat_violation_values(worksheet: Worksheet, dataframe: DataFrame, body_val_list: list = None,
                                    workbook: Workbook = None, full_stat_path: str = None) -> bool:
    """Заполнение акта значениями из dataframe

    :param full_stat_path:
    :param workbook:
    :param body_val_list:
    :param worksheet: страница для заполнения
    :param dataframe: dataframe с данными нарушений
    :return: bool
    """
    serial_number = 0
    clean_categories: list = await db_get_categories_list()

    for category in clean_categories:

        # TODO  переделать под подсчет для статистики
        violation_value = await get_violation_value(dataframe, category)

        if not violation_value:
            continue

        serial_number += 1
        for body_val_item in body_val_list:

            if category.get('title') != body_val_item.get('value'):
                continue

            negative = len([i for i in violation_value if i.get('status', None) == 0])
            in_work = len([i for i in violation_value if i.get('status', None) == 2])
            in_act = len([i for i in violation_value if i.get('in_act', None)])

            violation_dict = {
                'v_len': f'{len(violation_value)} \\ {in_act}',
                'v_negative': f'{negative} \\ {in_work}',
            }
            await set_worksheet_cell_value(worksheet, body_val_item, serial_number, violation_dict)
            workbook.save(full_stat_path)
            break

    for body_val_item in body_val_list:

        if body_val_item.get('value') == "В С Е Г О":
            negative = len(dataframe[dataframe['status_id'] == 0])
            in_work = len(dataframe[dataframe['status_id'] == 2])
            in_act = len(dataframe[dataframe['act_number'] != ""])

            violation_dict_all = {
                'v_len': f'{len(dataframe.index)} \\ {in_act}',
                'v_negative': f'{negative} \\ {in_work}',
            }
            await set_worksheet_cell_value(worksheet, body_val_item, serial_number + 1, violation_dict_all)
            workbook.save(full_stat_path)
            break

    return True


async def set_page_setup(worksheet):
    """Установка параметров страницы

    :param worksheet:
    :return:
    """

    #  https://xlsxwriter.readthedocs.io/page_setup.html
    # worksheet.print_title_rows = '$2:$3'
    # worksheet.print_title = '$2:$3'

    # Printer Settings
    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4

    # Подогнать область печати к определенному кол-у страниц как по вертикали, так и по горизонтали.
    worksheet.page_setup.fitToPage = True
    worksheet.page_setup.fitToHeight = '0'

    # worksheet.views
    worksheet.print_options.horizontalCentered = True
    worksheet.print_area = '$A$1:$I$48'

    #  масштабный коэффициент для распечатываемой страницы
    # worksheet.set_print_scale(75)

    # worksheet.row_breaks.append(Break(id=45))
    worksheet.col_breaks.append(Break(id=9))


STATISTIC_RANGE_COLUMNS_BASIC = [
    ["A", "3"],
    ["B", "15.22"],
    ["C", "41"],
    ["D", "6"],
    ["E", "46.5"],
    ["F", "6"],
    ["G", "28"],
    ["H", "28"],
]

STATISTIC_MERGED_CELLS_HEADER = [
    'D2:H2',
    'D3:E3', 'F3:H3',
    'C4:H4',
    'D5:H5',
    'D6:H6',
    'D7:H7',
    'D8:H8',
    'D9:H9',
    'D10:E10', 'F10:H10',
    'F11:H11',
    'F12:H12',
    'F13:H13',
    'F14:H14',
    'F15:H15',
    'C11:C16',
    'F16:H16',
    'C17:H17',
    'D18:E18',
]

# # STATISTIC_MERGED_CELLS_BODY = [
# #     'C19:C42',
# # ]
#
# # STATISTIC_MERGED_CELLS_FOOTER = [
# #     'C43:H43',
# #     'C44:H44',
# # ]
#
STATISTIC_ROW_DIMENSIONS_HEADER = [
    ["2", "39.0"],
    ["3", "25.8"],
    ["4", "17.4"],
    ["5", "17.4"],
    ["6", "17.4"],
    ["7", "17.4"],
    ["8", "17.4"],
    ["9", "17.4"],
    ["10", "17.4"],
    ["11", "24.0"],
    ["12", "24.0"],
    ["13", "24.0"],
    ["14", "24.0"],
    ["15", "24.0"],
    ["16", "24.0"],
    ["17", "17.4"],
    ["18", "40"],
]

# STATISTIC_ROW_DIMENSIONS_BODY = [
#     ["19", "33.0"],
#     ["20", "33.0"],
#     ["21", "33.0"],
#     ["22", "33.0"],
#     ["23", "33.0"],
#     ["24", "33.0"],
#     ["25", "33.0"],
#     ["26", "33.0"],
#     ["27", "33.0"],
#     ["28", "33.0"],
#     ["29", "33.0"],
#     ["30", "33.0"],
#     ["31", "33.0"],
#     ["32", "33.0"],
#     ["33", "33.0"],
#     ["34", "33.0"],
#     ["35", "33.0"],
#     ["36", "33.0"],
#     ["37", "33.0"],
#     ["38", "33.0"],
#     ["39", "33.0"],
#     ["40", "33.0"],
#     ["41", "33.0"],
# ]
#
# STATISTIC_ROW_DIMENSIONS_FOOTER = [
#     ["42", "33.0"],
#     ["43", "17.4"],
#     ["44", "58.2"],
#     ["45", "33.0"],
#     ["46", "33.0"],
# ]
STATISTIC_CELL_RANGE_BACKGROUND_COLOR_HEADER = [
    ["C2:H3", "FF7030A0"],
    ["C4:H4", "FFFFC000"],
    ["C5:C16", "FFDCDAFA"],
    ["D10:H10", "FFDCDAFA"],
    ["C17:H17", "FFFFC000"],
    ["C18:H18", "FFDCDAFA"],
]
# STATISTIC_CELL_RANGE_BACKGROUND_COLOR_BODY = [
#     ["C19:C42", "FFDCDAFA"],
# ]
# STATISTIC_CELL_RANGE_BACKGROUND_COLOR_FOOTER = [
#     ["C43:H43", "FFFFC000"],
# ]
STATISTIC_CELL_RANGES_HEADER = [
    ('C2:H18', False),
]
#
# STATISTIC_CELL_RANGES_BODY = [
#     ('C19:H44', False),
# ]
#
# STATISTIC_CELL_RANGES_FOOTER = [
#     ('C2:H44', False),
# ]
#
STATISTIC_CELL_RANGES_BASIC_ALIGNMENT = [
    ('C2:H44', 'Times New Roman', 14, 'center', 'center'),
]

STATISTIC_CELL_RANGES_ALIGNMENT_HEADER = [
    ("C2:C2", 'Times New Roman', 14, 'center', 'center'),
    ("C3:H3", 'Times New Roman', 14, 'center', 'center'),
    ("C4:H4", 'Times New Roman', 14, 'center', 'center'),
    ("C10:H10", 'Times New Roman', 14, 'center', 'center'),
    ("C17:H17", 'Times New Roman', 14, 'center', 'center'),
    ("C18:H18", 'Times New Roman', 14, 'center', 'center'),

]
STATISTIC_CELL_RANGES_ALIGNMENT_BODY = [

]
# STATISTIC_CELL_RANGES_ALIGNMENT_FOOTER = [
#     ("C43:H43", 'Times New Roman', 14, 'center', 'center'),
#     ("C44:H44", 'Times New Roman', 14, 'center', 'center'),
# ]

STATISTIC_CELL_RANGES_SET_REPORT_FONT_HEADER = [
    ["C2:H3", {"color": "FFFFFFFF", "font_size": "14", "bold": "True", "name": "Arial"}],
    ["C2:C2", {"color": "FFFFFFFF", "font_size": "16", "bold": "True", "name": "Arial"}],
    ["C4:H4", {"font_size": "14", "bold": "True", "name": "Arial"}],
    ["C5:H10", {"font_size": "14", "bold": "True", "name": "Arial"}],
    ["D10:H10", {"font_size": "14", "bold": "True", "italic": "True", "name": "Arial"}],
    ["D17:H17", {"font_size": "14", "bold": "True", "name": "Arial"}],
    ["C18:C18", {"font_size": "14", "bold": "True", "name": "Arial"}],
    ["D18:H18", {"font_size": "14", "bold": "True", "italic": "True", "name": "Arial"}],
]
STATISTIC_CELL_RANGES_SET_REPORT_FONT_BODY = [

]
STATISTIC_CELL_RANGES_SET_REPORT_FONT_FOOTER = [
    ["C43:H44", {"font_size": "14", "bold": "True", "name": "Arial"}],
]


async def set_worksheet_cell_value(worksheet: Worksheet, item: dict, serial_number: int, violation_dict: dict) -> bool:
    """Формирование тела отчета (раздел итогов)

    :param violation_dict: dict
    :param worksheet: страница для заполнения
    :param item:
    :param serial_number:
    :return:
    """
    # worksheet.cell(row=int(item['row']), column=6, value=serial_number)
    worksheet.cell(row=int(item['row']), column=7, value=violation_dict.get('v_len', ''))
    worksheet.cell(row=int(item['row']), column=8, value=violation_dict.get('v_negative', ''))

    cell_ranges = [
        [f"D{item['row']}:F{item['row']}", {"color": "FF0000", "font_size": "14", "name": "Arial"}],
        [f"G{item['row']}:G{item['row']}", {"color": "FF0000", "font_size": "12", "name": "Arial"}],
        [f"H{item['row']}:H{item['row']}", {"color": "FF0000", "font_size": "12", "name": "Arial"}],
    ]
    for cell_range in cell_ranges:
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])
        await set_stat_alignment(worksheet, cell_range[0], horizontal='center', vertical='center')

    # worksheet.row_dimensions[int(item['row'])].height = await get_height_for_row(
    #     sheet=worksheet, row_number=int(item['row']),
    #     value="".join([str(i["description"]) + ' \n' for i in val]))

    return True


async def sets_report_font(worksheet, cell_range, params: dict) -> bool:
    """Форматирование ячейки: размер шрифта

    """
    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.font = Font(
                color=params.get("color"),
                italic=params.get("italic"),
                size=params.get("font_size"),
                bold=params.get("bold"),
                name=params.get("name"),
                underline=params.get("underline"),
                vertAlign=params.get("vertAlign"),
            )

        except Exception as err:
            logger.error(f"item {item} cell {cell}")
            logger.error(f"sets_report_font {repr(err)}")
            continue
    return True


async def set_stat_alignment(worksheet, cell_range, horizontal=None, vertical=None):
    """Форматирование ячейки: положение текста в ячейке (лево верх)
    """
    wrap_alignment = Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.alignment = wrap_alignment

        except Exception as err:
            logger.error(f"iter {item} cell {cell}")
            logger.error(f"set_mip_alignment {repr(err)}")
            continue
    return True


async def set_column_dimensions(worksheet, column, width):
    """Форматирование: ширина столбцов

    :param worksheet:
    :param column:
    :param width:
    :return:
    """
    try:
        worksheet.column_dimensions[column].width = width

    except Exception as err:
        logger.error(f"set_column_widths {repr(err)}")


async def set_column_dimensions(worksheet, column, width):
    """Форматирование: ширина столбцов

    :param worksheet:
    :param column:
    :param width:
    :return:
    """
    try:
        worksheet.column_dimensions[column].width = width

    except Exception as err:
        logger.error(f"set_column_widths {repr(err)}")


async def set_report_font(worksheet, cell_range, font_size=14, font_name='Arial') -> bool:
    """Форматирование ячейки: шрифт и размер шрифта

    """
    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.font = Font(size=font_size, name=font_name)

        except Exception as err:
            logger.error(f"item {item} cell {cell}")
            logger.error(f"set_report_font {repr(err)}")
            continue
    return True


async def set_merge_cells(worksheet: Worksheet, merged_cell: str) -> bool:
    """Форматирование: обьединение ячеек

    :param worksheet:
    :param merged_cell:
    :return:
    """
    try:
        worksheet.merge_cells(merged_cell)
        return True

    except Exception as err:
        logger.error(f"set_merge_cells {repr(err)}")
        return False


async def set_range_border(worksheet, cell_range, border=None):
    """Форматирование ячейки: все границы ячейки

    border - только нижняя граница
    """

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    if border:
        thin_border = Border(bottom=Side(style='thin'))

    rows = worksheet[cell_range]

    try:
        for row in rows:
            for cell in row:
                try:
                    cell.border = thin_border

                except Exception as err:
                    logger.error(f"set_border {repr(err)}")
                    continue

    except Exception as err:
        logger.error(f"set_border {repr(err)}")
        return False


async def set_row_dimensions(worksheet: Worksheet, row_number: int or str, height: int or str) -> bool:
    """Установление высоты строки

    :param worksheet:
    :param row_number:
    :param height:
    :return:
    """
    try:
        worksheet.row_dimensions[int(row_number)].height = float(height)
        return True

    except Exception as err:
        logger.error(f"set_row_dimensions {repr(err)}")
        return False


async def set_report_background_color(worksheet, cell_range, rgb='FF7030A0') -> bool:
    """Установка цвета заливки ячейки
    :return:
    """
    rows = worksheet[cell_range]
    for row in rows:
        for cell in row:
            try:
                color_fill = PatternFill(start_color=rgb,
                                         end_color=rgb,
                                         fill_type='solid')
                cell.fill = color_fill
                return True

            except Exception as err:
                logger.error(f"set_border {repr(err)}")
                return False


async def get_violation_value(dataframe: DataFrame, category: dict) -> list:
    violation_value: list = []

    category_id = category.get('id', None)

    if category_id is None:
        return []

    category_dataframe: DataFrame = dataframe.loc[dataframe['category_id'] == category_id]
    if category_dataframe.empty:
        return []

    for row_df in category_dataframe.itertuples(index=False):
        try:
            description = row_df.description.value[0]
            elimination_time_id = row_df.elimination_time_id
            status_id = row_df.status_id
            in_act = row_df.act_number

            violation_value.append(
                {
                    "description": description,
                    "elimination_time": elimination_time_id,
                    "status": status_id,
                    'in_act': in_act,
                }
            )

        except Exception as err:
            logger.error(f"{repr(err)}")
            continue

    return violation_value
