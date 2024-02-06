from __future__ import annotations

import os
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.page import PageMargins
from pandas import DataFrame

from apps.MyBot import bot_send_message
from apps.core.bot.handlers.generate.generate_support_paths import (get_image_name,
                                                                    get_dop_photo_full_filename)
from apps.core.bot.messages.messages import Messages
from apps.core.bot.reports.report_data import headlines_data
from apps.core.database.query_constructor import QueryConstructor
from apps.core.database.db_utils import (db_get_data_dict_from_table_with_id,
                                         db_get_data_list,
                                         db_get_clean_headers)
from config.config import Udocan_media_path

from loader import logger


async def create_xlsx(chat_id: int | str, full_act_path: str) -> (Workbook, Worksheet):
    """Создание файла xlsx с нуля

    :param: chat_id
    :param: full_act_path

    """

    is_created: bool = await create_new_xlsx(report_file=full_act_path)
    if is_created is None:
        logger.warning(Messages.Error.workbook_not_create)
        await bot_send_message(chat_id=chat_id, text=Messages.Error.workbook_not_create)
        return None, None

    workbook: Workbook = await get_workbook(fill_report_path=full_act_path)
    if workbook is None:
        logger.warning(Messages.Error.workbook_not_found)
        await bot_send_message(chat_id=chat_id, text=Messages.Error.workbook_not_found)
        return None, None

    worksheet: Worksheet = await get_worksheet(workbook, index=0)
    if worksheet is None:
        logger.warning(Messages.Error.worksheet_not_found)
        await bot_send_message(chat_id=chat_id, text=Messages.Error.worksheet_not_found)
        return None, None

    return workbook, worksheet


async def create_new_xlsx(report_file: str) -> bool:
    """Создание xlsx
    """
    try:
        wb = openpyxl.Workbook()
        wb.save(report_file)
        return True

    except Exception as err:
        logger.error(F"set_border {repr(err)}")
        return False


async def get_worksheet(wb: Workbook, index: int = 0) -> Worksheet | None:
    """Получение Страницы из документа по индексу
    :param wb: Workbook - книга xls
    :param index: int - индекс листа
    :return: worksheet or None
    """
    try:
        worksheet: Worksheet = wb.worksheets[index]
        return worksheet

    except Exception as err:
        logger.error(f"get_workbook {repr(err)}")
        return None


async def get_workbook(fill_report_path: str) -> Workbook | None:
    """Открыть и загрузить Workbook
    :param fill_report_path: str полный путь к файлу
    :return: Workbook or None
    """
    try:
        workbook: Workbook = openpyxl.load_workbook(fill_report_path)
        return workbook

    except Exception as err:
        logger.error(f"get_workbook {repr(err)}")
        return None


async def anchor_photo(dataframe, row_number: int, workbook: Workbook, worksheet: Worksheet, full_act_path: str,
                       act_date: str = None):
    """Вставка фото в ячейку с учетом смещения. Возвращает область печати с учетом фото

    """
    photo_row: int = 63 + row_number
    row_num: int = 0
    img_params: dict = {"height": 220,
                        "anchor": True,
                        "scale": True, }

    row_act_header: int = photo_row - 1
    row_act_photo: int = photo_row
    img_params["row_header"] = row_act_header
    img_params["row_description"] = row_act_photo
    img_params["row"] = row_act_photo

    table_name: str = 'core_violations'
    clean_headers: list = await db_get_clean_headers(table_name=table_name)

    for num_data, violation_data in enumerate(dataframe.itertuples(index=False), start=1):
        # for num_data, violation_data in enumerate(df.to_dict('index'), start=1):

        violation_dict = dict(zip(clean_headers, violation_data))

        img_params["photo_full_name"] = violation_dict.get('photo_full_name', None)
        if num_data % 2 != 0:
            img_params["column"] = 'B'
            img_params["column_img"] = 2
            img_params["column_description"] = 6
            img_params["description"] = violation_dict.get('description', None)

            if num_data != 1:
                img_params["row_header"] = img_params["row_header"] + 2
                img_params["row_description"] = img_params["row_description"] + 2
                img_params["row"] = img_params["row"] + 2
                await format_act_photo_header(worksheet, img_params["row_header"])
                await format_act_photo_description(worksheet, img_params["row_description"])
                workbook.save(full_act_path)
                photo_row += 2
        else:
            img_params["column"] = 'H'
            img_params["column_img"] = 8
            img_params["column_description"] = 12
            img_params["description"] = violation_dict.get('description', None)

        is_anchor: bool = False
        try:
            is_anchor = await set_act_photographic_materials(worksheet, img_params)

        except Exception as err:
            logger.error(f"set_act_photographic_materials {num_data= } {row_num=} {repr(err)} ")

        if is_anchor:
            workbook.save(full_act_path)
            # set header
            worksheet.cell(row=img_params["row_header"],
                           column=img_params["column_img"],
                           value=f'Фото {num_data} к пункту {num_data} ')
            workbook.save(full_act_path)

            # set description
            worksheet.cell(row=img_params["row_description"],
                           column=img_params["column_description"],
                           value=img_params["description"])
            workbook.save(full_act_path)

    print_area = f'$A$1:M{photo_row + row_num + 1}'

    return print_area, photo_row + row_num + 1


async def anchor_dop_photo(chat_id, dataframe: DataFrame, start_row_number: int, workbook: Workbook,
                           worksheet: Worksheet, full_act_path: str) -> str:
    """Добавление дополнительных фотоматериалов"""

    img_params: list = await get_img_params(
        user_id=chat_id, dataframe=dataframe, start_row_num=start_row_number
    )

    photo_row: int = start_row_number
    for photo_num, img_param in enumerate(img_params):
        try:
            await set_act_photographic_materials(worksheet, img_param)
            await set_act_photographic_materials_description(worksheet, img_param, photo_num)
            photo_row: int = img_param.get('row')
            workbook.save(full_act_path)

        except Exception as err:
            logger.error(f"set_act_photographic_materials {photo_num= } {repr(err)} ")
            continue

    print_area = f'$A$1:M{photo_row + 1}'
    await set_act_page_after_footer_setup(worksheet, print_area)

    return print_area


async def get_img_params(user_id: str, dataframe: DataFrame, start_row_num: int) -> list:
    """Получение параметров изображений

    :return:
    """
    group_photo_list: list = await get_group_photo_list(user_id=user_id, dataframe=dataframe)

    photo_row: int = start_row_num

    row_act_header: int = photo_row - 1
    row_act_photo: int = photo_row

    img_params: dict = {
        "height": 220,
        "anchor": True,
        "scale": True,
        "row_header": row_act_header,
        "row_description": row_act_photo,
        "row": row_act_photo
    }

    img_params_list: list = []
    for num_data, media_file_dict in enumerate(group_photo_list, start=1):

        img_params["photo_full_name"] = media_file_dict.get('photo_full_name', None)

        if num_data % 2 != 0:
            img_params["column"] = 'B'
            img_params["column_img"] = 2
            img_params["column_description"] = 6
            img_params["description"] = media_file_dict.get('description', None)

            if num_data != 1:
                img_params["row_header"] = img_params["row_header"] + 2
                img_params["row_description"] = img_params["row_description"] + 2
                img_params["row"] = img_params["row"] + 2
                photo_row += 2
        else:
            img_params["column"] = 'H'
            img_params["column_img"] = 8
            img_params["column_description"] = 12
            img_params["description"] = media_file_dict.get('description', None)

        img_params_list.append(img_params)

    return img_params_list


async def get_group_photo_list(user_id: str, dataframe: DataFrame) -> list:
    """

    :return:
    """
    clean_headers: list = await db_get_clean_headers(table_name='core_violations')

    group_list: list = []
    for num_data, violation_data in enumerate(dataframe.itertuples(index=False), start=1):
        violation_dict: dict = dict(zip(clean_headers, violation_data))
        if not violation_dict.get('media_group', None): continue
        media_group_list: list = violation_dict.get('media_group', None).split(':::')

        media_group: list = []
        for item in media_group_list:
            photo_full_name: str = await get_dop_photo_full_filename(user_id=user_id, name=item)
            media_group.append(
                {
                    'photo_full_name': photo_full_name,
                    'dop_photo': item
                }
            )
        group_list.append(
            {
                'num_data': num_data,
                'media_group_list': media_group_list
            }
        )

    return group_list


async def insert_service_image(worksheet: Worksheet, *, chat_id: int = None, service_image_name: str = 'Logo',
                               img_params: dict = None) -> bool:
    """Вставка изображений в файл

    :param service_image_name: str - имя файла для обработки
    :param chat_id: int - id пользователя (папки) где находится logo
    :param img_params: dict параметры вставки
    :param worksheet: Worksheet - объект страницы файла
    :return: bool
    """

    photo_full_name: str = await get_image_name(Udocan_media_path, "HSE", str(chat_id), f"{service_image_name}.jpg")

    # if chat_id:
    #     photo_full_name: str = await get_image_name(Udocan_media_path, "HSE",
    #     str(chat_id), f"{service_image_name}.jpg")

    if not os.path.isfile(photo_full_name):
        logger.error("service image not found")
        photo_full_name: str = await get_image_name(Udocan_media_path, "HSE", str(chat_id), "Logo.jpg")

    if not img_params:
        img_params: dict = {
            'photo_full_name': photo_full_name,
            "height": 90,
            "width": 230,
            "anchor": True,
            "column": 'B',
            "column_img": 2,
            "row": 2,
        }

    if not os.path.isfile(photo_full_name):
        logger.error("service image not found")
        return False

    img: Image = Image(photo_full_name)
    img = await image_preparation(img, img_params)

    result = await insert_images(worksheet, img=img)
    return bool(result)


async def image_preparation(img: Image, img_params: dict) -> Image:
    """Подготовка изображения перед вставкой на страницу

    :param: img - объект изображения
    :param: img_params - параметры изображения
    :return: Image - измененный объект изображения
    """
    height = img_params.get("height")
    scale = img_params.get("scale")
    width = img_params.get("width")
    anchor = img_params.get("anchor")
    column = img_params.get("column")
    row = img_params.get("row")

    # высота изображения
    if height:
        img.height = height

    # ширина изображения
    if width:
        img.width = width

    # изменение пропорций изображения
    if scale:
        scale = img.height / max(img.height, img.width)
        img.width = img.width * scale

    # прикрепление изображение по адресу str(column) + str(row)
    if anchor:
        img.anchor = str(column) + str(row)

    return img


async def insert_images(worksheet: Worksheet, img: Image) -> bool:
    """Вставка изображения на лист worksheet,

    :param img: файл изображения
    :param worksheet: Worksheet - страница документа для вставки изображения
    :return: bool
    """

    try:
        worksheet.add_image(img)
        return True

    except Exception as err:
        logger.error(f"insert_images {repr(err)}")
        return False


async def set_act_body_values(worksheet: Worksheet):
    """Вставка информации на лист отчета

    :param worksheet: Worksheet - объект страницы файла
    :return:
    """
    values = [
        {'coordinate': 'B6', 'value': 'Акт предписание № от', 'row': '6', 'column': '2'},
        {'coordinate': 'B7', 'value': 'об устранении нарушений ', 'row': '7', 'column': '2'},

        {'coordinate': 'B9', 'value': 'тут наименование организации', 'row': '9', 'column': '2'},
        {'coordinate': 'B10',
         'value': '(указать кому адресовано (полное или сокращенное наименование юридического '
                  'лица либо индивидуального предпринимателя, ИНН) ',
         'row': '10', 'column': '2'},
        {'coordinate': 'B11', 'value': 'Мною:', 'row': '11', 'column': '2'},
        {'coordinate': 'B12', 'value': 'тут должность и ФИО полностью выдавшего', 'row': '12', 'column': '2'},

        {'coordinate': 'B14',
         'value': '(фамилия, имя, отчество (последнее – при наличии), должность должностного '
                  'лица, уполномоченного выдавать предписания',
         'row': '14', 'column': '2'},
        {'coordinate': 'B15', 'value': 'проведена проверка по соблюдению требований ОТ, ПБ и ООС, в отношении:',
         'row': '15', 'column': '2'},
        {'coordinate': 'B16', 'value': 'тут наименование организации', 'row': '16', 'column': '2'},
        {'coordinate': 'B17',
         'value': '(указать полное наименование юридического лица либо индивидуального предпринимателя)', 'row': '17',
         'column': '2'},

        {'coordinate': 'B19', 'value': 'В присутствии:', 'row': '19', 'column': '2'},
        {'coordinate': 'B20', 'value': 'тут ответственный', 'row': '20', 'column': '2'},

        {'coordinate': 'B22', 'value': 'ПРЕДПИСЫВАЕТСЯ ', 'row': '22', 'column': '2'},
        {'coordinate': 'B23', 'value': 'Принять меры по устранению выявленных нарушений в установленные сроки.',
         'row': '23', 'column': '2'},

        {'coordinate': 'B25', 'value': '№', 'row': '25', 'column': '2'},
        {'coordinate': 'B25', 'value': 'Описание и характер выявленных нарушений', 'row': '25', 'column': '3'},
        {'coordinate': 'B25',
         'value': 'Наименование НПА, номера подпунктов, пунктов, требования которых нарушены или не соблюдены',
         'row': '25', 'column': '6'},
        {'coordinate': 'B25', 'value': 'Предписываемые меры по устранению выявленного нарушения', 'row': '25',
         'column': '9'}, {'coordinate': 'B25', 'value': 'Срок устранения нарушений', 'row': '25', 'column': '12'},
        {'coordinate': 'B26', 'value': 'п/п', 'row': '26', 'column': '2'},
        {'coordinate': 'B27', 'value': 'тут наименование ПО', 'row': '27', 'column': '2'},
        {'coordinate': 'B28', 'value': '1', 'row': '28', 'column': '2'},
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column'])).value = str(val['value'])

        except Exception as err:
            logger.error(f"set_values {repr(err)}")
            return None


async def set_act_page_after_footer_setup(worksheet: Worksheet, print_area: str, break_line: int = None) -> bool:
    """Установка параметров страницы

    :param break_line: int
    :param print_area: str
    :param worksheet: Worksheet - объект страницы файла
    :return: bool
    """

    #  https://xlsxons.horizontalCentered = True
    worksheet.print_area = print_area

    #  масштабный коэффициент для распечатываемой страницы
    # worksheet.set_print_scale(75)

    if break_line:
        worksheet.row_breaks.append(Break(id=break_line))
        # worksheet.col_breaks.append(Break(id=13))

    return True


async def set_act_violation_values(worksheet: Worksheet, dataframe: DataFrame, workbook: Workbook,
                                   full_act_path: str) -> int:
    """Заполнение акта значениями из dataframe. Сортировка по main_location_id и sub_locations_id

    :param full_act_path:
    :param workbook:
    :param worksheet: Worksheet - объект страницы файла для заполнения
    :param dataframe: dataframe с данными нарушений
    :return: int
    """
    # берём уникальные значения main_location_id
    unique_main_locations_ids: list = dataframe.main_location_id.unique().tolist()
    row_number: int = 0
    item_row_value: int = 0

    for main_location_id in unique_main_locations_ids:
        new_header_row = True

        # разделяем dataframe по каждому уникальному значению
        main_location_values: dataframe = dataframe.loc[dataframe['main_location_id'] == main_location_id]
        if main_location_values.empty:
            continue

        # разделяем по sub_location
        unique_sub_locations_ids: list = main_location_values.sub_location_id.unique().tolist()

        for sub_locations_id in unique_sub_locations_ids:

            new_header_row = True
            sub_locations_values: dataframe = main_location_values.loc[dataframe['sub_location_id'] == sub_locations_id]
            if sub_locations_values.empty:
                continue

            # итерируемся по dataframe как по tuple
            for row in sub_locations_values.itertuples(index=False):
                item_row_value += 1
                row_number = await set_act_worksheet_cell_value(
                    worksheet, row, row_number, new_header_row, workbook, full_act_path, item_row_value
                )
                new_header_row = False
                row_number += 1

    return row_number


async def set_act_worksheet_cell_value(worksheet: Worksheet, violation_values: DataFrame, row_number: int,
                                       new_header_row: bool, workbook: Workbook, full_act_path: str,
                                       item_row_value: int) -> int:
    """Заполнение тела акта каждым пунктом

    :param item_row_value:
    :param full_act_path:
    :param workbook:
    :param new_header_row: bool  Требуется ли новый заголовок?
    :param row_number:
    :param violation_values: DataFrame
    :param worksheet: Worksheet - объект страницы файла для заполнения
    :return
    """

    if row_number == 0:
        logger.debug(violation_values)

        await set_single_violation(worksheet, violation_values)
        workbook.save(full_act_path)
        return row_number

    if new_header_row:
        logger.debug(f'{new_header_row= }')

        await set_value_title(worksheet, violation_values, row_number)
        workbook.save(full_act_path)
        row_number += 1

        await set_act_violation(worksheet, violation_values, row_number=row_number, item_row_value=item_row_value)
        workbook.save(full_act_path)
        return row_number

    if not new_header_row:
        logger.debug(f'{new_header_row= }')

        await set_act_violation(worksheet, violation_values, row_number=row_number, item_row_value=item_row_value)
        workbook.save(full_act_path)
        return row_number


async def set_act_violation(worksheet: Worksheet, violation_values: DataFrame, row_number: int,
                            item_row_value: int) -> int:
    """Ставка данных нарушения в строку

    :param item_row_value: - номер строки
    :param row_number: - номер строки
    :param worksheet: Worksheet - объект страницы файла
    :param violation_values: DataFrame - с данными нарушения
    :return: int - номер заполненной строки
    """
    row_value: int = 28 + row_number

    worksheet.cell(row=row_value, column=2, value=item_row_value)
    normative_document: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_normativedocuments',
        post_id=violation_values.normative_documents_id
    )

    if violation_values.description:
        worksheet.cell(row=row_value, column=3, value=violation_values.description)
        # worksheet.cell(row=row_value, column=15, value=violation_values.description)
    else:
        worksheet.cell(row=row_value, column=3, value=normative_document['title'])
        # worksheet.cell(row=row_value, column=15, value=normative_document['title'])

    worksheet.cell(row=row_value, column=6, value=normative_document['normative'])
    # worksheet.cell(row=row_value, column=16, value=normative_document['normative'])

    if violation_values.comment and violation_values.comment not in [None, '', ' ', '.', '*', '/']:
        worksheet.cell(row=row_value, column=9, value=violation_values.comment)
        # worksheet.cell(row=row_value, column=17, value=violation_values.comment)
    else:
        worksheet.cell(row=row_value, column=9, value=normative_document['procedure'])
        # worksheet.cell(row=row_value, column=17, value=normative_document['procedure'])

    elimination_time: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_eliminationtime',
        post_id=violation_values.elimination_time_id
    )
    elimination_data = (datetime.datetime.strptime(violation_values.created_at, '%Y-%m-%d')
                        + datetime.timedelta(days=elimination_time['days'])).strftime('%d.%m.%Y')

    worksheet.cell(row=row_value, column=12, value=elimination_data)

    merged_cell = [
        f'C{row_value}:E{row_value}',
        f'F{row_value}:H{row_value}',
        f'I{row_value}:K{row_value}',
    ]

    for item in merged_cell:
        await set_merge_cells(worksheet, merged_cell=item)

    act_range_border = [
        (f'B{row_value}:L{row_value}', False)
    ]
    for item in act_range_border:
        await set_range_border(worksheet, cell_range=item[0], border=item[1])

    row_dimensions = [
        [f'{row_value}', '105'],
    ]
    for item in row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    report_font = [
        (f'B{row_value}:L{row_value}', 'Times New Roman', 11, 'center', 'center'),
        (f'O{row_value}:Q{row_value}', 'Times New Roman', 11, 'center', 'center')
    ]

    elimination_time: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_eliminationtime',
        post_id=violation_values.elimination_time_id
    )
    elimination_data = (datetime.datetime.strptime(violation_values.created_at, '%Y-%m-%d')
                        + datetime.timedelta(days=elimination_time['days'])).strftime('%d.%m.%Y')

    worksheet.cell(row=row_value, column=12, value=elimination_data)

    for item_range in report_font:
        await set_report_font(
            worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1]
        )

    for item_range in report_font:
        await set_act_alignment(
            worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4]
        )

    await set_automatic_row_dimensions(
        worksheet, row_number=row_value, row_value=violation_values
    )
    return row_value


async def set_single_violation(worksheet: Worksheet, violation_values: DataFrame):
    """Заполнение акта из единственного пункта

    :param
    """
    main_location: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_mainlocation',
        post_id=violation_values.main_location_id
    )
    sub_location: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_sublocation',
        post_id=violation_values.sub_location_id
    )

    title: str = f"{main_location['title']} ({sub_location['title']})"

    worksheet.cell(row=27, column=2, value=title)
    worksheet.cell(row=28, column=2, value=1)

    normative_document: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_normativedocuments',
        post_id=violation_values.normative_documents_id
    )

    if violation_values.description:
        worksheet.cell(row=28, column=3, value=violation_values.description)
        # worksheet.cell(row=28, column=15, value=violation_values.description)
    else:
        worksheet.cell(row=28, column=3, value=normative_document['title'])
        # worksheet.cell(row=28, column=15, value=normative_document['title'])

    worksheet.cell(row=28, column=6, value=normative_document['normative'])
    # worksheet.cell(row=28, column=16, value=normative_document['normative'])

    if violation_values.comment and violation_values.comment not in ['', ' ', '.', '*', '/']:
        worksheet.cell(row=28, column=9, value=violation_values.comment)
        # worksheet.cell(row=28, column=17, value=violation_values.comment)
    else:
        worksheet.cell(row=28, column=9, value=normative_document['procedure'])
        # worksheet.cell(row=28, column=17, value=normative_document['procedure'])

    elimination_time: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_eliminationtime',
        post_id=violation_values.elimination_time_id
    )
    elimination_data = (datetime.datetime.strptime(violation_values.created_at, '%Y-%m-%d')
                        + datetime.timedelta(days=elimination_time['days'])).strftime('%d.%m.%Y')

    await set_automatic_row_dimensions(worksheet, row_number=28, row_value=violation_values)

    worksheet.cell(row=28, column=12, value=elimination_data)


async def set_automatic_row_dimensions(worksheet: Worksheet, row_number: int, row_value) -> bool:
    """Автоматическое установление высоты строки по тексту

    :param row_value:
    :param worksheet: Worksheet - объект страницы файла
    :param row_number:
    :return:
    """
    if not row_value:
        return False

    normative_documents: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_normativedocuments',
        post_id=row_value.normative_documents_id)

    title = normative_documents.get('title', None)
    normative = normative_documents.get('normative', None)
    procedure = normative_documents.get('procedure', None)
    comment = row_value.comment

    if not normative:
        logger.error(f"No normative found if {row_value =  }")
    if not procedure:
        logger.error(f"No procedure found if {row_value =  }")

    if procedure:
        comment = procedure if len(row_value.comment) < len(procedure) else row_value.comment

    item_list: list = []
    try:
        list_val = [row_value.description, title, normative, comment, procedure]
        for item in list_val:
            if not isinstance(item, str): continue

            lines = min(len(item.split("\n\n")) - 1, 1)
            height = max(len(item) / 26 + lines, 1.5) * 16
            item_list.append(height)

        max_height = round(max(item_list), 2) + 15

        max_height = max_height if max_height > 60 else 60

        dim = worksheet.row_dimensions[row_number]
        dim.height = max_height

        logger.debug(f"row_number {row_number} max_height {max_height}")
        return True

    except Exception as err:
        logger.error(f"Error row {row_number} set_automatic_row_dimensions {repr(err)}")
        return False


async def set_act_alignment(worksheet, cell_range, horizontal=None, vertical=None):
    """Форматирование ячейки: положение текста в ячейке (лево верх)

    """
    wrap_alignment = Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.alignment = wrap_alignment

        except Exception as err:
            logger.error(f"iter {item} cell {cell}")
            logger.error(f"set_mip_alignment {repr(err)}")


async def set_report_font(worksheet, cell_range, font_size=14, font_name='Arial') -> bool:
    """Форматирование ячейки: шрифт и размер шрифта

    """
    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.font = Font(size=font_size, name=font_name)
        except Exception as err:
            logger.error(f"item {item} cell {cell}")
            logger.error(f"set_report_font {repr(err)}")
            continue
    return True


async def set_value_title(worksheet: Worksheet, violation_values: DataFrame, row_number: int) -> bool:
    """

    :param row_number:
    :param violation_values:
    :param worksheet: Worksheet - объект страницы файла
    :return:
    """
    row_value = 28 + row_number
    main_location: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_mainlocation',
        post_id=violation_values.main_location_id
    )
    sub_location: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_sublocation',
        post_id=violation_values.sub_location_id
    )

    title: str = f"{main_location['short_title']} ({sub_location['title']})"
    worksheet.cell(row=row_value, column=2, value=title)

    merged_cell = [
        f'B{row_value}:L{row_value}',
    ]
    for item in merged_cell:
        await set_merge_cells(
            worksheet, merged_cell=item
        )

    act_range_border = [
        (f'B{row_value}:L{row_value}', False)
    ]
    for item in act_range_border:
        await set_range_border(
            worksheet, cell_range=item[0], border=item[1]
        )

    row_dimensions = [
        [f'{row_value}', '18'],
    ]
    for item in row_dimensions:
        await set_row_dimensions(
            worksheet, row_number=item[0], height=item[1]
        )

    report_font = [
        (f'B{row_value}:L{row_value}', 'Times New Roman', 14, 'center', 'center'),
    ]
    for item_range in report_font:
        await set_report_font(
            worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1]
        )

    for item_range in report_font:
        await set_act_alignment(
            worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4]
        )

    report_font = [
        [f'B{row_value}:L{row_value}',
         {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
    ]

    for cell_range in report_font:
        await sets_report_font(
            worksheet, cell_range[0], params=cell_range[1]
        )

    return True


async def sets_report_font(worksheet, cell_range, params: dict) -> bool:
    """Форматирование ячейки: размер шрифта

    """
    cells = [cell for row in worksheet[cell_range] for cell in row]

    for item, cell in enumerate(cells, start=1):
        try:
            cell.font = Font(
                color=params.get("color"),
                italic=params.get("italic"),
                size=params.get("font_size"),
                bold=params.get("bold"),
                name=params.get("name"),
                underline=params.get("underline"),
                vertAlign=params.get("vertAlign"),
            )

        except Exception as err:
            logger.error(f"item {item} cell {cell}")
            logger.error(f"sets_report_font {repr(err)}")
            continue
    return True


async def set_row_dimensions(worksheet: Worksheet, row_number: int | str, height: int):
    """Установление высоты строки

    :param worksheet: Worksheet - объект страницы файла
    :param row_number: int | str - номер страницы
    :param height: int - высота
    :return:
    """
    try:
        worksheet.row_dimensions[int(row_number)].height = float(height)
    except Exception as err:
        logger.error(f"set_row_dimensions {repr(err)}")


async def set_range_border(worksheet, cell_range, border=None):
    """Форматирование ячейки: все границы ячейки

    border - только нижняя граница
    """

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    if border:
        thin_border = Border(bottom=Side(style='thin'))

    rows = worksheet[cell_range]

    try:
        for row in rows:
            for cell in row:
                try:
                    cell.border = thin_border

                except Exception as err:
                    logger.error(f"set_border {repr(err)}")
                    continue

    except Exception as err:
        logger.error(f"set_border {repr(err)}")
        return False


async def set_merge_cells(worksheet: Worksheet, merged_cell: str) -> bool:
    """Форматирование: обьединение ячеек

    :param worksheet: Worksheet - объект страницы файла
    :param merged_cell:
    :return:
    """
    try:
        worksheet.merge_cells(merged_cell)
        return True
    except Exception as err:
        logger.error(f"set_merge_cells {repr(err)}")
        return False


async def set_act_photographic_materials(worksheet: Worksheet, img_params: dict) -> bool:
    """Вставка фото нарушения на страницу акта

    :param img_params:
    :param worksheet: Worksheet - объект страницы файла
    :return:
    """
    if not img_params:
        logger.error(f'set_act_photographic_materials not found dict img_params {img_params = }')
        return False

    photo_full_name = img_params.get("photo_full_name", None)
    if not photo_full_name:
        logger.error(
            f'set_act_photographic_materials not found param img_params["photo_full_name"] {photo_full_name = }')
        return False

    if not os.path.isfile(photo_full_name):
        logger.error(f'set_act_photographic_materials photo not found {photo_full_name}')
        return False

    img: Image = Image(photo_full_name)
    img = await image_preparation(img, img_params)
    await insert_images(worksheet, img=img)

    return True


async def set_act_photographic_materials_description(worksheet, img_params, photo_num):
    # set header
    worksheet.cell(
        row=img_params["row_header"], column=img_params["column_img"], value=f'Фото {photo_num} к пункту {photo_num}'
    )
    # set description
    worksheet.cell(
        row=img_params["row_description"], column=img_params["column_description"],
        value=img_params["description"]
    )


async def format_act_photo_description(worksheet: Worksheet, row_number: int):
    """Форматирование строк с фото материалами

    :param row_number:
    :param worksheet: Worksheet - объект страницы файла
    :return:
    """
    merged_cells = [

        f'B{row_number}:E{row_number}',
        f'H{row_number}:K{row_number}',
    ]
    for merged_cell in merged_cells:
        await set_merge_cells(worksheet, merged_cell)

    photographic_materials_alignment = [
        f'B{row_number}:L{row_number}',
    ]
    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_act_alignment(worksheet, cell_range, horizontal='center', vertical='center')

    # for item, cell_range in enumerate(photographic_materials_alignment, start=1):
    #     await set_range_border(worksheet, cell_range=cell_range)

    photographic_row_dimensions = [
        [f'{row_number}', "166"],
    ]
    for item in photographic_row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_report_font(worksheet, cell_range=cell_range, font_size=14)

    photographic_report_font = [
        [f'B{row_number}:L{row_number}', {"font_size": "9", "name": "Arial"}],
    ]
    for item, cell_range in enumerate(photographic_report_font, start=1):
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

    # background_color = [
    #     ["C50:H50", "FFFFC000"]
    # ]
    # for item in background_color:
    #     await set_report_background_color(worksheet, item[0], rgb=item[1])


async def format_act_photo_header(worksheet: Worksheet, row_number: int) -> bool:
    """Форматирование строк с фото материалами

    :param row_number:
    :param worksheet: Worksheet - объект страницы файла
    :return: bool
    """
    merged_cells = [
        f'B{row_number}:E{row_number}',
        f'H{row_number}:K{row_number}',
    ]
    for merged_cell in merged_cells:
        await set_merge_cells(worksheet, merged_cell)

    photographic_materials_alignment = [
        f'B{row_number}:L{row_number}',
    ]
    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_act_alignment(worksheet, cell_range, horizontal='center', vertical='center')

    # for item, cell_range in enumerate(photographic_materials_alignment, start=1):
    #     await set_range_border(worksheet, cell_range=cell_range)

    photographic_row_dimensions = [
        [f'{row_number}', "18"],
    ]
    for item in photographic_row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_report_font(worksheet, cell_range=cell_range, font_size=14)

    photographic_report_font = [
        [f'B{row_number}:L{row_number}', {"font_size": "12", "bold": "True", "name": "Arial"}],
    ]
    for item, cell_range in enumerate(photographic_report_font, start=1):
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

    return True


async def set_act_photographic_materials_values(worksheet: Worksheet, row_value: int):
    """

    :param row_value:
    :param worksheet: Worksheet - объект страницы файла
    """
    values = [
        {'coordinate': 'K59', 'row': f'{59 + row_value}', 'column': '12', 'value': 'Приложение 1', },
        {'coordinate': 'K60', 'row': f'{60 + row_value}', 'column': '12', 'value': 'к 1 части Акта-Предписания', },
        {'coordinate': 'B62', 'row': f'{62 + row_value}', 'column': '2', 'value': 'Фото 1 к пункту 1', },
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column']), value=str(val['value']))

        except Exception as err:
            logger.error(f"set_act_photographic_materials_values {repr(err)}")
            continue


async def set_act_footer_values(worksheet, row_number):
    """

     :param row_number:
     :param worksheet: Worksheet - объект страницы файла
     :return:
     """
    row_value = 28 + row_number

    values = [
        {'coordinate': 'B30',
         'value': 'Информацию о выполнении пунктов настоящего предписания необходимо направить '
                  f'в письменной форме в адрес {headlines_data.get("hse_organization_full_name", None)} '
                  f'по адресу: {headlines_data.get("hse_organization_legal_address", None)} или ',
         'row': f'{30 - 28 + row_value}', 'column': '2'},
        {'coordinate': 'B31', 'value': 'на эл. адреса:',
         'row': f'{31 - 28 + row_value}', 'column': '2'},
        {'coordinate': 'G31', 'value': 'тут эл.адрес выдавшего',
         'row': f'{31 - 28 + row_value}', 'column': '7'},
        {'coordinate': 'B32', 'value': 'не позднее:',
         'row': f'{32 - 28 + row_value}', 'column': '2'},
        {'coordinate': 'G32', 'value': 'тут крайняя дата ответа',
         'row': f'{32 - 28 + row_value}', 'column': '7'},
        {'coordinate': 'B33',
         'value': 'Невыполнение предписания в установленный срок является основанием '
                  'для применения дисциплинарных взысканий',
         'row': f'{33 - 28 + row_value}', 'column': '2'},

        {'coordinate': 'B35', 'value': 'С предписанием ознакомлен: ',
         'row': f'{35 - 28 + row_value}', 'column': '2'},
        {'coordinate': 'B36',
         'value': '(Ф.И.О, подпись должность руководителя, иного должностного лица или уполномоченного представителя '
                  'юридического лица, индивидуального предпринимателя, его уполномоченного представителя)',
         'row': f'{36 - 28 + row_value}', 'column': '2'},

        {'coordinate': 'B38', 'value': 'Сведения об отказе в ознакомлении с предписанием и отказе от подписи',
         'row': f'{38 - 28 + row_value}', 'column': '2'},

        {'coordinate': 'B40',
         'value': '(фамилия, имя, отчество (последнее – при наличии), проводившего(их) проверку или '
                  'уполномоченного выдавать предписания)',
         'row': f'{40 - 28 + row_value}', 'column': '2'},

        {'coordinate': 'B42', 'value': 'Предписание выдал: ',
         'row': f'{42 - 28 + row_value}', 'column': '2'},
        {'coordinate': 'B43', 'value': 'тут должность выдавшего',
         'row': f'{43 - 28 + row_value}', 'column': '2'},
        {'coordinate': 'B43', 'value': 'тут ФИО выдавшего',
         'row': f'{43 - 28 + row_value}', 'column': '11'},
        {'coordinate': 'B44', 'value': '(должность)',
         'row': f'{44 - 28 + row_value}', 'column': '2'},
        {'coordinate': 'B44', 'value': '(подпись)',
         'row': f'{44 - 28 + row_value}', 'column': '8'},
        {'coordinate': 'B44', 'value': '(расшифровка подписи)',
         'row': f'{44 - 28 + row_value}', 'column': '11'},

        {'coordinate': 'B46', 'value': 'тут дата',
         'row': f'{46 - 28 + row_value}', 'column': '11'},
        {'coordinate': 'B47', 'value': '(дата выдачи)',
         'row': f'{47 - 28 + row_value}', 'column': '11'},

        {'coordinate': 'B49', 'value': 'Экземпляр предписания для исполнения получил:',
         'row': f'{49 - 28 + row_value}', 'column': '2'},

        {'coordinate': 'B51',
         'value': '(фамилия, имя, отчество (последнее – при наличии), должность руководителя, '
                  'иного должностного лица или уполномоченного представителя юридического лица, '
                  'индивидуального предпринимателя, его уполномоченного представителя)',
         'row': f'{51 - 28 + row_value}', 'column': '2'},

        {'coordinate': 'B53', 'value': '(дата)',
         'row': f'{53 - 28 + row_value}', 'column': '11'},

        {'coordinate': 'B55', 'value': '(подпись)',
         'row': f'{55 - 28 + row_value}', 'column': '11'},
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column'])).value = str(val['value'])

        except Exception as err:
            logger.error(f"set_values {repr(err)}")
            return None


async def set_act_header_values(worksheet, headlines_data: dict = None) -> bool:
    """Заполнение заголовка отчета

    :param headlines_data:
    :param worksheet: Worksheet - объект страницы файла
    :return:
    """
    values: list = [
        {"coordinate": "B6",
         "value": f"Акт предписание № {headlines_data.get('act_number', 'act_number')} от "
                  f"{headlines_data.get('act_date', 'act_date')} г.",
         "row": "6",
         "column": "2"},

        {"coordinate": "B9",
         "value": f"{headlines_data.get('general_contractor_full_name', 'general_contractor_full_name')}",
         "row": "9",
         "column": "2"},

        {"coordinate": "B12",
         "value": f"{headlines_data.get('HSE_function_dative', 'HSE_function_dative')} "
                  f"{headlines_data.get('hse_department_dative', 'hse_department_dative')} "
                  f"{headlines_data.get('HSE_full_name_dative', 'HSE_full_name_dative')}",
         "row": "12",
         "column": "2"},

        {"coordinate": "B16",
         "value": f"{headlines_data.get('general_contractor_full_name', 'general_contractor_full_name')}",
         "row": "16",
         "column": "2"},

        {"coordinate": "B20",
         "value": f"{headlines_data.get('contractor_representative', 'contractor_representative')}",
         "row": "20",
         "column": "2"},
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column']), value=str(val['value']))
        except Exception as err:
            logger.error(f"set_user_values {repr(err)}")
            continue

    return True


async def set_act_footer_footer_values(worksheet: Worksheet, row_number):
    """Установка значений футера акта

    :param worksheet: Worksheet - объект страницы файла
    :param row_number:
    :return:
    """
    row_value = 28 + row_number

    values: list = [
        {"coordinate": "G31",
         "value": f"{headlines_data.get('HSE_email')}",
         "row": f"{31 - 28 + row_value}",
         "column": "7"},

        {"coordinate": "G32",
         "value": f"{(headlines_data.get('act_max_date'))}",
         "row": f"{32 - 28 + row_value}",
         "column": "7"},

        {"coordinate": "B43",
         "value": f"{headlines_data.get('HSE_function')} {headlines_data.get('hse_department_dative')}",
         "row": f"{43 - 28 + row_value}",
         "column": "2"},

        {"coordinate": "B43",
         "value": f"{headlines_data.get('HSE_name_short')}",
         "row": f"{43 - 28 + row_value}",
         "column": "11"},

        {"coordinate": "B46",
         "value": f"{headlines_data.get('act_date')}",
         "row": f"{46 - 28 + row_value}",
         "column": "11"},
    ]

    for val in values:
        try:
            worksheet.cell(row=int(val['row']), column=int(val['column']), value=str(val['value']))
        except Exception as err:
            logger.error(f"set_user_values {repr(err)}")
            continue


async def get_act_headlines_data_values(chat_id, dataframe: DataFrame = None, act_date: str = None,
                                        act_number: int = None) -> dict:
    """ Формирование данных для формирования шапки акта - предписания

    :return: dict headlines_data
    """

    # if not headlines_data:
    table_name: str = 'core_violations'
    clean_headers: list = await db_get_clean_headers(table_name=table_name)
    query_kwargs: dict = {
        "action": 'SELECT', "subject": '*',
        "conditions": {
            "hse_telegram_id": chat_id,
        },
    }
    query: str = await QueryConstructor(None, 'core_hseuser', **query_kwargs).prepare_data()
    datas_query_list: list = await db_get_data_list(query=query)
    clean_datas_query_list = datas_query_list[0]

    item_dict: dict = dict(
        (header, item_value) for header, item_value in zip(clean_headers, clean_datas_query_list))

    hse_id = item_dict.get('id', None)

    if not hse_id:
        logger.error(f'hse_id is {hse_id}')
        return {}

    query_kwargs: dict = {
        "action": 'SELECT', "subject": '*',
        "conditions": {
            "hse_telegram_id": chat_id,
        },
    }
    query: str = await QueryConstructor(None, 'core_hseuser', **query_kwargs).prepare_data()
    # print(f'{__name__} {fanc_name()} {query}')

    hse_user: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_hseuser',
        post_id=hse_id,
        query=query
    )

    # ФИО выдавшего
    # headlines_data['HSE'] = list(set(list(dataframe.name)))[0]
    headlines_data['HSE'] = hse_user.get('hse_full_name', None)

    # ФИО сокращенно выдавшего
    # headlines_data['HSE_name_short'] = 'Коршаков А.С.'
    headlines_data['HSE_name_short'] = hse_user.get('hse_short_name', None)

    # ФИО и должность выдавшего в родительном
    # headlines_data['HSE_full_name_dative'] = list(set(list(dataframe.name)))[0]
    headlines_data['HSE_full_name_dative'] = hse_user.get('hse_full_name_dative', None)

    # Email выдавшего
    # headlines_data['HSE_email'] = 'as.korshakov@udokancopper.com'
    headlines_data['HSE_email'] = hse_user.get('hse_contact_main_email', None)

    # Должность выдавшего
    # headlines_data['HSE_function'] = list(set(list(dataframe.function)))[0]
    headlines_data['HSE_function'] = hse_user.get('hse_function', None)
    headlines_data['HSE_function_dative'] = hse_user.get('hse_function_dative', None)

    # Отдел выдавшего
    # headlines_data['hse_department'] = list(set(list(dataframe.function)))[0]
    headlines_data['hse_department'] = hse_user.get('hse_department', None)
    headlines_data['hse_department_dative'] = hse_user.get('hse_department_dative', None)

    headlines_data['act_number'] = act_number if act_number else 00000

    # Максимальная дата ответа
    elimination_time_id = max(dataframe['elimination_time_id'])
    elimination_time: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_eliminationtime',
        post_id=elimination_time_id
    )

    max_date = max(dataframe['created_at'])
    days_max = elimination_time['days']
    headlines_data['act_date'] = datetime.datetime.strptime(max_date, '%Y-%m-%d').strftime('%d.%m.%Y')
    headlines_data['act_max_date'] = (datetime.datetime.strptime(act_date, '%d.%m.%Y') + datetime.timedelta(
        days=days_max)).strftime('%d.%m.%Y')

    # Подрядчик
    general_contractor_id = list(set(list(dataframe.general_contractor_id)))[0]
    general_contractor_dict: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_generalcontractor',
        post_id=general_contractor_id)

    headlines_data['general_contractor_full_name'] = general_contractor_dict.get('title', None)
    headlines_data['general_contractor_legal_address'] = general_contractor_dict.get('legal_address', None)

    hse_organization: dict = await db_get_data_dict_from_table_with_id(
        table_name='core_generalcontractor',
        post_id=hse_user.get('hse_organization', None))

    headlines_data['hse_organization_full_name'] = hse_organization.get('title', None)
    headlines_data['hse_organization_legal_address'] = hse_organization.get('legal_address', None)

    # TODO  update general_contractor table in DB
    # Ответственное лицо подрядчика
    headlines_data['contractor_representative'] = ' '

    # Проект
    main_location_id_list: list = list(set(list(dataframe.main_location_id)))
    item_dict: dict = {}
    for main_location_id in main_location_id_list:
        main_location_name: dict = await db_get_data_dict_from_table_with_id(
            table_name='core_mainlocation',
            post_id=main_location_id)
        item_dict[main_location_id] = main_location_name

    headlines_data['main_location'] = item_dict

    headlines_data["day"] = (datetime.datetime.now()).strftime("%d")
    headlines_data["year"] = (datetime.datetime.now()).strftime("%Y")

    return headlines_data


async def format_act_footer_prescription_sheet(worksheet: Worksheet, row_number: int, act_num: str = None) -> bool:
    """Пошаговое форматирование страницы

    :param act_num:
    :param row_number:
    :param worksheet: Worksheet - объект страницы файла
    :return: bool
    """
    await set_act_page_setup(worksheet, act_num)

    ACT_FOOTER_MERGED_CELLS = [
        f'B{30 + row_number}:L{30 + row_number}', f'B{31 + row_number}:F{31 + row_number}',
        f'G{31 + row_number}:L{31 + row_number}', f'B{32 + row_number}:F{32 + row_number}',
        f'G{32 + row_number}:L{32 + row_number}', f'B{33 + row_number}:L{33 + row_number}',
        f'B{35 + row_number}:L{35 + row_number}', f'B{36 + row_number}:L{36 + row_number}',
        f'B{38 + row_number}:L{38 + row_number}', f'B{39 + row_number}:L{39 + row_number}',
        f'B{40 + row_number}:L{40 + row_number}', f'B{42 + row_number}:F{42 + row_number}',
        f'B{43 + row_number}:F{43 + row_number}', f'H{43 + row_number}:I{43 + row_number}',
        f'K{43 + row_number}:L{43 + row_number}', f'B{44 + row_number}:F{44 + row_number}',
        f'H{44 + row_number}:I{44 + row_number}', f'K{44 + row_number}:L{44 + row_number}',
        f'K{46 + row_number}:L{46 + row_number}', f'K{47 + row_number}:L{47 + row_number}',
        f'B{49 + row_number}:L{49 + row_number}', f'B{50 + row_number}:L{50 + row_number}',
        f'B{51 + row_number}:L{51 + row_number}', f'K{52 + row_number}:L{52 + row_number}',
        f'K{53 + row_number}:L{53 + row_number}', f'K{54 + row_number}:L{54 + row_number}',
        f'K{55 + row_number}:L{55 + row_number}',
    ]
    for merged_cell in ACT_FOOTER_MERGED_CELLS:
        await set_merge_cells(worksheet, merged_cell=merged_cell)

    ACT_FOOTER_CELL_RANGES = [
        (f'B{35 + row_number}:L{35 + row_number}', False),
        (f'B{36 + row_number}:L{36 + row_number}', False),
        (f'B{38 + row_number}:L{38 + row_number}', False),
        (f'B{39 + row_number}:L{39 + row_number}', False),
        (f'B{40 + row_number}:L{40 + row_number}', False),
        (f'B{43 + row_number}:F{43 + row_number}', True),
        (f'H{43 + row_number}:I{43 + row_number}', True),
        (f'K{43 + row_number}:L{43 + row_number}', True),
        (f'K{46 + row_number}:L{46 + row_number}', True),
        (f'B{50 + row_number}:L{50 + row_number}', True),
        (f'K{52 + row_number}:L{52 + row_number}', True),
        (f'K{54 + row_number}:L{54 + row_number}', True),
    ]
    for item in ACT_FOOTER_CELL_RANGES:
        await set_range_border(worksheet, cell_range=item[0], border=item[1])

    ACT_FOOTER_ROW_DIMENSIONS = [
        [f'{29 + row_number}', '5.5'], [f'{30 + row_number}', '45'], [f'{31 + row_number}', '18'],
        [f'{32 + row_number}', '18'], [f'{33 + row_number}', '30'], [f'{34 + row_number}', '5.5'],
        [f'{35 + row_number}', '18'], [f'{36 + row_number}', '32'], [f'{37 + row_number}', '5.5'],
        [f'{38 + row_number}', '20'], [f'{39 + row_number}', '18'], [f'{40 + row_number}', '30'],
        [f'{41 + row_number}', '18'], [f'{42 + row_number}', '18'], [f'{43 + row_number}', '30'],
        [f'{44 + row_number}', '18'], [f'{45 + row_number}', '18'], [f'{46 + row_number}', '18'],
        [f'{47 + row_number}', '18'], [f'{48 + row_number}', '18'], [f'{49 + row_number}', '18'],
        [f'{50 + row_number}', '30'], [f'{51 + row_number}', '40'], [f'{52 + row_number}', '18'],
        [f'{53 + row_number}', '18'], [f'{54 + row_number}', '18'], [f'{55 + row_number}', '18'],
    ]
    for item in ACT_FOOTER_ROW_DIMENSIONS:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    ACT_FOOTER_CELL_RANGES_BASIC_ALIGNMENT = [
        (f'B{30 + row_number}:L{30 + row_number}', 'Times New Roman', 11, 'center', 'center'),
        (f'B{31 + row_number}:L{31 + row_number}', 'Times New Roman', 11, 'center', 'center'),
        (f'B{32 + row_number}:L{32 + row_number}', 'Times New Roman', 11, 'center', 'center'),
        (f'B{33 + row_number}:L{33 + row_number}', 'Times New Roman', 12, 'center', 'center'),
        (f'B{35 + row_number}:L{35 + row_number}', 'Times New Roman', 14, 'left', 'center'),
        (f'B{36 + row_number}:L{36 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{38 + row_number}:L{38 + row_number}', 'Times New Roman', 12, 'center', 'center'),
        (f'B{40 + row_number}:L{40 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{42 + row_number}:L{42 + row_number}', 'Times New Roman', 14, 'center', 'center'),
        (f'B{43 + row_number}:L{43 + row_number}', 'Times New Roman', 12, 'center', 'center'),
        (f'B{44 + row_number}:L{44 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{46 + row_number}:L{46 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{47 + row_number}:L{47 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{49 + row_number}:L{49 + row_number}', 'Times New Roman', 12, 'left', 'center'),
        (f'B{51 + row_number}:L{51 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{53 + row_number}:L{53 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{55 + row_number}:L{55 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{59 + row_number}:L{59 + row_number}', 'Times New Roman', 10, 'right', 'center'),
        (f'B{60 + row_number}:L{60 + row_number}', 'Times New Roman', 10, 'right', 'center'),
    ]
    for item_range in ACT_FOOTER_CELL_RANGES_BASIC_ALIGNMENT:
        await set_report_font(worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1])

    for item_range in ACT_FOOTER_CELL_RANGES_BASIC_ALIGNMENT:
        await set_act_alignment(worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4])

    print_area = f'$A$1:M{56 + row_number}'
    break_line = 58 + row_number
    await set_act_page_after_footer_setup(worksheet, print_area, break_line)

    return True


async def format_act_prescription_sheet(worksheet: Worksheet, act_num: int):
    """Пошаговое форматирование страницы
    """
    await set_act_page_setup(worksheet, act_num)

    for item in ACT_RANGE_COLUMNS:
        await set_column_dimensions(worksheet, column=item[0], width=item[1])

    for merged_cell in ACT_MERGED_CELLS:
        await set_merge_cells(worksheet, merged_cell=merged_cell)

    for item in ACT_CELL_RANGES:
        await set_range_border(worksheet, cell_range=item[0], border=item[1])

    for item in ACT_ROW_DIMENSIONS:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    for item_range in ACT_CELL_RANGES_BASIC_ALIGNMENT:
        await set_report_font(worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1])

    for item_range in ACT_CELL_RANGES_BASIC_ALIGNMENT:
        await set_act_alignment(worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4])

    for cell_range in ACT_CELL_RANGES_SET_REPORT_FONT:
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])


async def format_act_footer_prescription_sheet(worksheet: Worksheet, row_number: int, act_number: int) -> bool:
    """Пошаговое форматирование страницы

    :param act_number:
    :param row_number:
    :param worksheet: Worksheet - объект страницы файла
    :return: bool
    """
    await set_act_page_setup(worksheet, act_num=act_number)

    ACT_FOOTER_MERGED_CELLS = [
        f'B{30 + row_number}:L{30 + row_number}', f'B{31 + row_number}:F{31 + row_number}',
        f'G{31 + row_number}:L{31 + row_number}', f'B{32 + row_number}:F{32 + row_number}',
        f'G{32 + row_number}:L{32 + row_number}', f'B{33 + row_number}:L{33 + row_number}',
        f'B{35 + row_number}:L{35 + row_number}', f'B{36 + row_number}:L{36 + row_number}',
        f'B{38 + row_number}:L{38 + row_number}', f'B{39 + row_number}:L{39 + row_number}',
        f'B{40 + row_number}:L{40 + row_number}', f'B{42 + row_number}:F{42 + row_number}',
        f'B{43 + row_number}:F{43 + row_number}', f'H{43 + row_number}:I{43 + row_number}',
        f'K{43 + row_number}:L{43 + row_number}', f'B{44 + row_number}:F{44 + row_number}',
        f'H{44 + row_number}:I{44 + row_number}', f'K{44 + row_number}:L{44 + row_number}',
        f'K{46 + row_number}:L{46 + row_number}', f'K{47 + row_number}:L{47 + row_number}',
        f'B{49 + row_number}:L{49 + row_number}', f'B{50 + row_number}:L{50 + row_number}',
        f'B{51 + row_number}:L{51 + row_number}', f'K{52 + row_number}:L{52 + row_number}',
        f'K{53 + row_number}:L{53 + row_number}', f'K{54 + row_number}:L{54 + row_number}',
        f'K{55 + row_number}:L{55 + row_number}',
    ]
    for merged_cell in ACT_FOOTER_MERGED_CELLS:
        await set_merge_cells(worksheet, merged_cell=merged_cell)

    ACT_FOOTER_CELL_RANGES = [
        (f'B{35 + row_number}:L{35 + row_number}', False),
        (f'B{36 + row_number}:L{36 + row_number}', False),
        (f'B{38 + row_number}:L{38 + row_number}', False),
        (f'B{39 + row_number}:L{39 + row_number}', False),
        (f'B{40 + row_number}:L{40 + row_number}', False),
        (f'B{43 + row_number}:F{43 + row_number}', True),
        (f'H{43 + row_number}:I{43 + row_number}', True),
        (f'K{43 + row_number}:L{43 + row_number}', True),
        (f'K{46 + row_number}:L{46 + row_number}', True),
        (f'B{50 + row_number}:L{50 + row_number}', True),
        (f'K{52 + row_number}:L{52 + row_number}', True),
        (f'K{54 + row_number}:L{54 + row_number}', True),
    ]
    for item in ACT_FOOTER_CELL_RANGES:
        await set_range_border(worksheet, cell_range=item[0], border=item[1])

    ACT_FOOTER_ROW_DIMENSIONS = [
        [f'{29 + row_number}', '5.5'], [f'{30 + row_number}', '45'], [f'{31 + row_number}', '18'],
        [f'{32 + row_number}', '18'], [f'{33 + row_number}', '30'], [f'{34 + row_number}', '5.5'],
        [f'{35 + row_number}', '18'], [f'{36 + row_number}', '32'], [f'{37 + row_number}', '5.5'],
        [f'{38 + row_number}', '20'], [f'{39 + row_number}', '18'], [f'{40 + row_number}', '30'],
        [f'{41 + row_number}', '18'], [f'{42 + row_number}', '18'], [f'{43 + row_number}', '30'],
        [f'{44 + row_number}', '18'], [f'{45 + row_number}', '18'], [f'{46 + row_number}', '18'],
        [f'{47 + row_number}', '18'], [f'{48 + row_number}', '18'], [f'{49 + row_number}', '18'],
        [f'{50 + row_number}', '30'], [f'{51 + row_number}', '40'], [f'{52 + row_number}', '18'],
        [f'{53 + row_number}', '18'], [f'{54 + row_number}', '18'], [f'{55 + row_number}', '18'],
    ]
    for item in ACT_FOOTER_ROW_DIMENSIONS:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    ACT_FOOTER_CELL_RANGES_BASIC_ALIGNMENT = [
        (f'B{30 + row_number}:L{30 + row_number}', 'Times New Roman', 11, 'center', 'center'),
        (f'B{31 + row_number}:L{31 + row_number}', 'Times New Roman', 11, 'center', 'center'),
        (f'B{32 + row_number}:L{32 + row_number}', 'Times New Roman', 11, 'center', 'center'),
        (f'B{33 + row_number}:L{33 + row_number}', 'Times New Roman', 12, 'center', 'center'),
        (f'B{35 + row_number}:L{35 + row_number}', 'Times New Roman', 14, 'left', 'center'),
        (f'B{36 + row_number}:L{36 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{38 + row_number}:L{38 + row_number}', 'Times New Roman', 12, 'center', 'center'),
        (f'B{40 + row_number}:L{40 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{42 + row_number}:L{42 + row_number}', 'Times New Roman', 14, 'center', 'center'),
        (f'B{43 + row_number}:L{43 + row_number}', 'Times New Roman', 12, 'center', 'center'),
        (f'B{44 + row_number}:L{44 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{46 + row_number}:L{46 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{47 + row_number}:L{47 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{49 + row_number}:L{49 + row_number}', 'Times New Roman', 12, 'left', 'center'),
        (f'B{51 + row_number}:L{51 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{53 + row_number}:L{53 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{55 + row_number}:L{55 + row_number}', 'Times New Roman', 10, 'center', 'center'),
        (f'B{59 + row_number}:L{59 + row_number}', 'Times New Roman', 10, 'right', 'center'),
        (f'B{60 + row_number}:L{60 + row_number}', 'Times New Roman', 10, 'right', 'center'),
    ]
    for item_range in ACT_FOOTER_CELL_RANGES_BASIC_ALIGNMENT:
        await set_report_font(worksheet, cell_range=item_range[0], font_size=item_range[2], font_name=item_range[1])

    for item_range in ACT_FOOTER_CELL_RANGES_BASIC_ALIGNMENT:
        await set_act_alignment(worksheet, item_range[0], horizontal=item_range[3], vertical=item_range[4])

    print_area = f'$A$1:M{56 + row_number}'
    break_line = 58 + row_number
    await set_act_page_after_footer_setup(worksheet, print_area, break_line)

    return True


async def format_act_photo_header(worksheet: Worksheet, row_number: int) -> bool:
    """Форматирование строк с фото материалами

    :param row_number:
    :param worksheet: Worksheet - объект страницы файла
    :return: bool
    """
    merged_cells = [
        f'B{row_number}:E{row_number}',
        f'H{row_number}:K{row_number}',
    ]
    for merged_cell in merged_cells:
        await set_merge_cells(worksheet, merged_cell)

    photographic_materials_alignment = [
        f'B{row_number}:L{row_number}',
    ]
    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_act_alignment(worksheet, cell_range, horizontal='center', vertical='center')

    # for item, cell_range in enumerate(photographic_materials_alignment, start=1):
    #     await set_range_border(worksheet, cell_range=cell_range)

    photographic_row_dimensions = [
        [f'{row_number}', "18"],
    ]
    for item in photographic_row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_report_font(worksheet, cell_range=cell_range, font_size=14)

    photographic_report_font = [
        [f'B{row_number}:L{row_number}', {"font_size": "12", "bold": "True", "name": "Arial"}],
    ]
    for item, cell_range in enumerate(photographic_report_font, start=1):
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

    return True


async def format_act_photo_description(worksheet: Worksheet, row_number: int) -> bool:
    """Форматирование строк с фото материалами

    :param row_number:
    :param worksheet: Worksheet - объект страницы файла
    :return:
    """
    merged_cells = [

        f'B{row_number}:E{row_number}',
        f'H{row_number}:K{row_number}',
    ]
    for merged_cell in merged_cells:
        await set_merge_cells(worksheet, merged_cell)

    photographic_materials_alignment = [
        f'B{row_number}:L{row_number}',
    ]
    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_act_alignment(worksheet, cell_range, horizontal='center', vertical='center')

    # for item, cell_range in enumerate(photographic_materials_alignment, start=1):
    #     await set_range_border(worksheet, cell_range=cell_range)

    photographic_row_dimensions = [
        [f'{row_number}', "166"],
    ]
    for item in photographic_row_dimensions:
        await set_row_dimensions(worksheet, row_number=item[0], height=item[1])

    for item, cell_range in enumerate(photographic_materials_alignment, start=1):
        await set_report_font(worksheet, cell_range=cell_range, font_size=14)

    photographic_report_font = [
        [f'B{row_number}:L{row_number}', {"font_size": "9", "name": "Arial"}],
    ]
    for item, cell_range in enumerate(photographic_report_font, start=1):
        await sets_report_font(worksheet, cell_range[0], params=cell_range[1])

    # background_color = [
    #     ["C50:H50", "FFFFC000"]
    # ]
    # for item in background_color:
    #     await set_report_background_color(worksheet, item[0], rgb=item[1])

    return True


async def set_column_dimensions(worksheet, column, width):
    """Форматирование: ширина столбцов

    :param worksheet: Worksheet - объект страницы файла
    :param column:
    :param width:
    :return:
    """
    try:
        worksheet.column_dimensions[column].width = width

    except Exception as err:
        logger.error(f"set_column_widths {repr(err)}")


async def set_act_page_setup(worksheet: Worksheet, act_num=None) -> bool:
    """Установка параметров страницы

    :param act_num:
    :param worksheet: Worksheet - объект страницы файла
    :return: bool
    """

    #  https://xlsxwriter.readthedocs.io/page_setup.html
    #  worksheet.print_title_rows = '$2:$3'
    #  worksheet.print_title = '$2:$3'

    # Printer Settings
    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4

    # Подогнать область печати к определенному кол-у страниц как по вертикали, так и по горизонтали.
    worksheet.page_setup.fitToPage = True
    worksheet.page_setup.fitToHeight = '0'

    # worksheet.views
    worksheet.print_options.horizontalCentered = True
    worksheet.print_area = '$A$1:M56'

    #  масштабный коэффициент для распечатываемой страницы
    #  worksheet.set_print_scale(75)

    #  worksheet.row_breaks.append(Break(id=53))
    #  worksheet.col_breaks.append(Break(id=13))

    # задаем собственные значения отступов
    top_bottom = 2 / 2.54
    left_right = 2 / 2.54
    worksheet.page_margins = PageMargins(left=left_right, right=left_right, top=top_bottom, bottom=top_bottom)

    worksheet.oddFooter.left.text = "Страница &[Page] из &N"

    if act_num:
        worksheet.oddFooter.left.text = f"Страница &[Page] из &N  акта № {act_num}"

    worksheet.oddFooter.left.size = 10
    worksheet.oddFooter.left.font = "Arial,Bold"
    worksheet.oddFooter.left.color = "030303"
    worksheet.differentFirst = False
    worksheet.differentOddEven = True

    return True


ACT_RANGE_COLUMNS = [
    ['A', '4'],
    ['B', '5'],
    ['C', '9'],
    ['D', '9'],
    ['E', '9'],
    ['F', '12'],
    ['G', '4'],
    ['H', '9'],
    ['I', '9'],
    ['J', '5'],
    ['K', '9'],
    ['L', '12'],
    ['M', '4'],

    ["O", "27"],
    ["P", "25"],
    ["Q", "23"],
]

ACT_MERGED_CELLS = [

    'B6:L6',
    'B7:L7',

    'B9:L9',
    'B10:L10',
    'B11:L11',
    'B12:L12',

    'B14:L14',
    'B15:L15',
    'B16:L16',
    'B17:L17',

    'B19:L19',
    'B20:L20',

    'B22:L22',
    'B23:L23',

    'C25:E26', 'F25:H26', 'I25:K26', 'L25:L26',

    'B27:L27',
    'C28:E28', 'F28:H28', 'I28:K28', 'L28:L28',
]

ACT_CELL_RANGES = [
    ('B9:L9', True),
    ('B12:L12', True),
    ('B16:L16', True),
    ('B20:L20', True),
    ('B25:E26', False), ('F25:H26', False), ('I25:K26', False), ('L25:L26', False),
    ('B27:L27', False),
    ('B28:E28', False), ('F28:H28', False), ('I28:K28', False), ('L28:L28', False),
]

ACT_ROW_DIMENSIONS = [
    ['2', '18'],
    ['3', '18'],
    ['4', '18'],
    ['5', '18'],
    ['6', '18'],
    ['7', '18'],
    ['8', '5.5'],
    ['9', '25'],
    ['10', '40'],
    ['11', '18'],
    ['12', '34'],
    ['13', '5.5'],
    ['14', '22'],
    ['15', '37'],
    ['16', '25'],
    ['17', '25'],
    ['18', '5.5'],
    ['19', '18'],
    ['20', '39'],
    ['21', '5.5'],
    ['22', '18'],
    ['23', '30'],
    ['24', '5.5'],
    ['25', '19'],
    ['26', '53'],
    ['27', '18'],
    ['28', '105'],

]

ACT_CELL_RANGES_BASIC_ALIGNMENT = [
    ('B6:L6', 'Times New Roman', 14, 'center', 'center'),
    ('B7:L7', 'Times New Roman', 14, 'center', 'center'),

    ('B9:L9', 'Verdana', 14, 'center', 'center'),
    ('B10:L10', 'Times New Roman', 10, 'center', 'center'),
    ('B11:L11', 'Times New Roman', 14, 'left', 'center'),
    ('B12:L12', 'Times New Roman', 14, 'left', 'center'),

    ('B14:L14', 'Times New Roman', 10, 'center', 'center'),
    ('B15:L15', 'Times New Roman', 14, 'center', 'center'),
    ('B16:L16', 'Verdana', 14, 'center', 'center'),
    ('B17:L17', 'Times New Roman', 10, 'center', 'center'),

    ('B19:L19', 'Times New Roman', 14, 'left', 'center'),
    ('B20:L20', 'Times New Roman', 14, 'left', 'center'),

    ('B22:L22', 'Times New Roman', 14, 'center', 'center'),
    ('B23:L23', 'Times New Roman', 14, 'center', 'center'),

    ('B25:L25', 'Times New Roman', 11, 'center', 'center'),
    ('B26:L26', 'Times New Roman', 11, 'center', 'center'),
    ('B27:L27', 'Times New Roman', 14, 'center', 'center'),
    ('B28:L28', 'Times New Roman', 11, 'center', 'center'),
    ('O28:Q28', 'Times New Roman', 11, 'center', 'center'),

]

ACT_CELL_RANGES_SET_REPORT_FONT = [
    ['B6:L6', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
    ['B7:L7', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
    ['B9:L9', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Verdana'}],
    ['B16:L16', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Verdana'}],
    ['B22:L22', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
    ['B27:L27', {'color': '000000', 'font_size': '14', 'bold': 'True', 'name': 'Times New Roman'}],
]
